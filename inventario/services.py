import re
import logging
import openpyxl
from datetime import datetime
from django.db import transaction
from django.utils import timezone

from .models import CicloInventario, ContaContabil, ItemInventario

logger = logging.getLogger(__name__)


class InventarioExcelImporter:
    """
    Serviço de importação e parsing de planilhas de Inventário Físico e Contábil (.xlsx)
    baseado no padrão oficial da PMESP / 2º BAEP.
    """

    def __init__(self, file_path_or_stream, titulo=None, ano=None, semestre=None, detentor=None, termo_numero=None, user=None):
        self.file_input = file_path_or_stream
        self.titulo = titulo or "Inventário Físico e Contábil de Material Permanente"
        self.ano = ano or timezone.now().year
        self.semestre = int(semestre) if semestre else 1
        self.detentor = detentor
        self.termo_numero = termo_numero
        self.user = user

    def process(self):
        wb = openpyxl.load_workbook(self.file_input, data_only=True)
        sheet = None
        
        # Procurar a aba principal do inventário
        for name in wb.sheetnames:
            if 'invent' in name.lower():
                sheet = wb[name]
                break
        if not sheet:
            sheet = wb.active

        # 1. Extração de metadados do cabeçalho
        extracted_termo = None
        extracted_detentor = None
        extracted_date = None

        for r in range(1, min(20, sheet.max_row + 1)):
            v = sheet.cell(r, 1).value
            if not v:
                continue
            str_v = str(v)
            if 'TERMO DE INVENTÁRIO' in str_v.upper() or 'TERMO DE INVENTARIO' in str_v.upper():
                extracted_termo = str_v.strip()
            elif 'Sr.' in str_v and 'Detentor Executivo' in str_v:
                m = re.search(r'Sr\.\s+([^,]+),', str_v)
                if m:
                    extracted_detentor = m.group(1).strip()
                m_date = re.search(r'Aos\s+(\d+)\s+dias do mês de\s+([a-zA-ZçÇ]+)\s+de\s+(\d{4})', str_v)
                if m_date:
                    extracted_date = m_date.group(0)

        termo_final = self.termo_numero or extracted_termo or f"2BAEP - INVENTÁRIO {self.semestre}S/{self.ano}"
        detentor_final = self.detentor or extracted_detentor or "Detentor Executivo 2º BAEP"

        with transaction.atomic():
            # 2. Criar o Ciclo de Inventário
            ciclo = CicloInventario.objects.create(
                titulo=self.titulo,
                termo_numero=termo_final,
                ano=self.ano,
                semestre=self.semestre,
                data_referencia=timezone.now().date(),
                detentor_executivo=detentor_final,
                status='EM_ANDAMENTO',
                observacoes=f"Importado via arquivo Excel em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                criado_por=self.user
            )

            current_account_code = None
            current_account_name = None
            current_account_obj = None
            contas_cache = {}

            itens_para_criar = []
            valor_total_acumulado = 0.0

            # 3. Iterar pelas linhas para extrair contas contábeis e bens
            for r in range(1, sheet.max_row + 1):
                c1 = sheet.cell(r, 1).value
                c2 = sheet.cell(r, 2).value

                # Identificação de cabeçalho de Conta Contábil (ex: "CONTA 123110301")
                if c1 and str(c1).strip().startswith('CONTA '):
                    raw_acc = str(c1).strip()
                    parts = raw_acc.split(' ', 1)
                    current_account_code = parts[1].strip() if len(parts) > 1 else raw_acc
                    
                    # Nome da conta normalmente está na célula da linha seguinte
                    acc_desc = sheet.cell(r+1, 1).value
                    current_account_name = str(acc_desc).strip() if acc_desc else f"CONTA {current_account_code}"

                    if current_account_code not in contas_cache:
                        conta_obj, _ = ContaContabil.objects.get_or_create(
                            codigo=current_account_code,
                            defaults={'nome': current_account_name}
                        )
                        contas_cache[current_account_code] = conta_obj
                    current_account_obj = contas_cache[current_account_code]
                    continue

                # Linha de Item: possui patrimônio numérico ou significativo na coluna 2
                if isinstance(c2, (int, float)) and c2 > 1000:
                    secao = str(c1 or '').strip()
                    patrimonio = str(int(c2) if isinstance(c2, float) and c2.is_integer() else c2).strip()
                    
                    c3 = sheet.cell(r, 3).value
                    n_serie = '' if c3 in (0, '0', None, 'NULL') else str(c3).strip()
                    
                    c4 = sheet.cell(r, 4).value
                    tipo_material = str(c4 or '').strip()
                    
                    c6 = sheet.cell(r, 6).value
                    situacao = str(c6 or '').strip() or 'EM USO'
                    
                    c8 = sheet.cell(r, 8).value
                    if isinstance(c8, (int, float)):
                        valor = float(c8)
                    else:
                        # Fallback se o valor estiver em outra coluna (ex: col 6)
                        try:
                            valor = float(str(c6).replace(',', '.'))
                        except Exception:
                            valor = 0.0

                    situacao_fisica = 'EM_EXCLUSAO' if 'EXCLUSÃO' in situacao.upper() else 'CONFORME'

                    item = ItemInventario(
                        ciclo=ciclo,
                        conta_contabil=current_account_obj,
                        secao_subunidade=secao,
                        patrimonio=patrimonio,
                        numero_serie=n_serie,
                        tipo_material=tipo_material,
                        situacao_material=situacao,
                        valor=valor,
                        conferido=False,
                        situacao_fisica_conferida=situacao_fisica
                    )
                    itens_para_criar.append(item)
                    valor_total_acumulado += valor

            # Bulk create para alta performance (1.000+ registros em milissegundos)
            ItemInventario.objects.bulk_create(itens_para_criar, batch_size=500)

            logger.info("Importação de inventário concluída. Ciclo ID %s, %s itens importados, R$ %.2f", ciclo.id, len(itens_para_criar), valor_total_acumulado)

            return {
                'ciclo': ciclo,
                'total_itens': len(itens_para_criar),
                'valor_total': valor_total_acumulado,
                'total_contas': len(contas_cache)
            }

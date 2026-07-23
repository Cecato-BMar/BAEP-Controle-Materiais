"""
Script para importação e sincronização completa da 'PLANILHA DE MATERIAS - Atualizada .xlsx'
Módulo Material Bélico — 2º BAEP
"""
import os
import sys
import datetime
import openpyxl
import django

# Configuração do Django
sys.path.insert(0, '/home/servidor-sys-baep/BAEP-Controle-Materiais')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'reserva_baep.settings')
django.setup()

from material_belico.models import (
    Fuzil, EspingardaCal12, PistolaGlock, PistolaTaurus, ArmaTransferenciaPendente,
    RedDot, Magnificador, Supressor, VinculacaoAcessorioFuzil,
    KitOperacional, RadioHT, AM640, AM600, MosquetaoFederal,
    TASER, Algemas, MunicaoQuimica, MunicaoConvencional,
    ColeteBalistico, EscudoBalistico, CapaceteBalistico
)

EXCEL_PATH = '/home/servidor-sys-baep/BAEP-Controle-Materiais/BAEP-Controle-Materiais-2/PLANILHA DE MATERIAS - Atualizada .xlsx'


def clean_val(v):
    if v is None:
        return None
    v_str = str(v).strip()
    if v_str in ('', 'None', '────────', '---------', 'N/I'):
        return None
    return v_str


def parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y'):
            try:
                return datetime.datetime.strptime(v, fmt).date()
            except ValueError:
                pass
    return None


def run_import():
    print(f"Carregando planilha oficial: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    resumo = {}

    # -------------------------------------------------------------------------
    # 1. FUZIS E ACESSÓRIOS
    # -------------------------------------------------------------------------
    if 'FUZIS E ACESSÓRIOS' in wb.sheetnames:
        ws = wb['FUZIS E ACESSÓRIOS']
        f_count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                # Tabela Esquerda: SCAR 762 (Col C=serie, D=loc, E=patrimonio, F=status)
                if len(row) >= 6 and row[2]:
                    serie_762 = clean_val(row[2])
                    loc_762 = clean_val(row[3]) or 'RESERVA'
                    pat_762 = clean_val(row[4]) or serie_762
                    st_762 = clean_val(row[5]) or 'OK'
                    if st_762 not in ['OK', 'EM_USO', 'BAIXADO', 'SINDICANCIA', 'MANUTENCAO']:
                        st_762 = 'OK'
                    if serie_762 and pat_762:
                        Fuzil.objects.update_or_create(
                            patrimonio=pat_762,
                            defaults={
                                'tipo': 'SCAR_762',
                                'localizacao': loc_762 if loc_762 in dict(Fuzil._meta.get_field('localizacao').choices) else 'RESERVA',
                                'status': st_762,
                                'observacoes': f'Série: {serie_762}'
                            }
                        )
                        f_count += 1
            except Exception as e:
                print(f"Erro em Fuzil 762: {e}")

            try:
                # Tabela Direita: SCAR/IA2 556 (Col I=serie, J=loc, K=patrimonio, L=status)
                if len(row) >= 12 and row[8]:
                    serie_556 = clean_val(row[8])
                    loc_556 = clean_val(row[9]) or 'RESERVA'
                    pat_556 = clean_val(row[10]) or serie_556
                    st_556 = clean_val(row[11]) or 'OK'
                    if st_556 not in ['OK', 'EM_USO', 'BAIXADO', 'SINDICANCIA', 'MANUTENCAO']:
                        st_556 = 'OK'
                    tipo_556 = 'IMBEL_IA2' if (serie_556 and serie_556.startswith('BRA')) else 'SCAR_556'
                    if serie_556 and pat_556:
                        Fuzil.objects.update_or_create(
                            patrimonio=pat_556,
                            defaults={
                                'tipo': tipo_556,
                                'localizacao': loc_556 if loc_556 in dict(Fuzil._meta.get_field('localizacao').choices) else 'RESERVA',
                                'status': st_556,
                                'observacoes': f'Série: {serie_556}'
                            }
                        )
                        f_count += 1
            except Exception as e:
                print(f"Erro em Fuzil 556: {e}")

        resumo['Fuzis'] = f_count

    # -------------------------------------------------------------------------
    # 2. PISTOLAS GLOCK
    # -------------------------------------------------------------------------
    if 'PISTOLAS GLOCK' in wb.sheetnames:
        ws = wb['PISTOLAS GLOCK']
        g_count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                if len(row) >= 4 and row[2]:
                    serie = clean_val(row[2])
                    patrimonio = clean_val(row[1])
                    modelo = clean_val(row[3]) or 'PISTOLA GLOCK G22 G5 .40'
                    cod_opm = clean_val(row[4])
                    unidade = clean_val(row[5]) or '2.BAEP'
                    situacao = clean_val(row[6]) or 'ok'
                    obs = clean_val(row[7])

                    st_map = {'OK': 'ok', 'EM USO': 'EM_USO', 'APREENDIDA': 'APREENDIDA', 'NOVIDADE': 'NOVIDADE'}
                    situacao_val = st_map.get(str(situacao).upper() if situacao else '', 'ok')

                    if serie:
                        PistolaGlock.objects.update_or_create(
                            numero_serie=serie,
                            defaults={
                                'patrimonio': patrimonio,
                                'modelo': modelo,
                                'cod_opm': cod_opm,
                                'unidade': unidade,
                                'situacao_reserva': situacao_val,
                                'observacoes': obs
                            }
                        )
                        g_count += 1
            except Exception as e:
                print(f"Erro em Pistola Glock: {e}")

        resumo['Pistolas Glock'] = g_count

    # -------------------------------------------------------------------------
    # 3. ESPINGARDAS CAL.12
    # -------------------------------------------------------------------------
    if 'CAL.12' in wb.sheetnames:
        ws = wb['CAL.12']
        e_count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                if len(row) >= 3 and row[2]:
                    num_esp = clean_val(row[2])
                    patrimonio = clean_val(row[3])
                    kit = clean_val(row[4])
                    status = clean_val(row[5]) or 'OK'
                    if num_esp:
                        EspingardaCal12.objects.update_or_create(
                            numero_espingarda=num_esp,
                            defaults={
                                'patrimonio': patrimonio,
                                'kit_vinculado': str(kit) if kit else None,
                                'status': status if status in ['OK', 'EM_USO', 'BAIXADO', 'MANUTENCAO'] else 'OK'
                            }
                        )
                        e_count += 1
            except Exception as e:
                print(f"Erro em Espingarda Cal.12: {e}")

        resumo['Espingardas Cal.12'] = e_count

    # -------------------------------------------------------------------------
    # 4. PISTOLAS TAURUS
    # -------------------------------------------------------------------------
    if 'PISTOLAS TAURUS' in wb.sheetnames:
        ws = wb['PISTOLAS TAURUS']
        t_count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                if len(row) >= 3 and row[2]:
                    serie = clean_val(row[2])
                    patrimonio = clean_val(row[1])
                    modelo_raw = clean_val(row[3]) or ''
                    unidade = clean_val(row[4]) or '2.BAEP'
                    obs = clean_val(row[5])

                    mod_code = 'TAURUS_24_7'
                    if 'PT100' in modelo_raw.upper():
                        mod_code = 'TAURUS_PT100'
                    elif '640' in modelo_raw.upper():
                        mod_code = 'TAURUS_640'

                    if serie:
                        PistolaTaurus.objects.update_or_create(
                            numero_serie=serie,
                            defaults={
                                'patrimonio': patrimonio,
                                'modelo': mod_code,
                                'unidade': unidade,
                                'observacoes': obs
                            }
                        )
                        t_count += 1
            except Exception as e:
                print(f"Erro em Pistola Taurus: {e}")

        resumo['Pistolas Taurus'] = t_count

    # -------------------------------------------------------------------------
    # 5. TRANSFERÊNCIAS PENDENTES (PST TRANSF)
    # -------------------------------------------------------------------------
    if 'PST TRANSF' in wb.sheetnames:
        ws = wb['PST TRANSF']
        tr_count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                if len(row) >= 7 and row[1]:
                    especie = clean_val(row[1])
                    marca = clean_val(row[2]) or 'N/I'
                    modelo = clean_val(row[3]) or 'N/I'
                    calibre = clean_val(row[4]) or 'N/I'
                    serie = clean_val(row[5]) or 'N/I'
                    nome = clean_val(row[6]) or 'N/I'
                    vinculo = clean_val(row[7]) or 'POLICIAL'
                    re_pol = clean_val(row[8])
                    situacao = clean_val(row[9]) or 'BAIXA'
                    status = clean_val(row[10]) or 'PARADO'
                    venda_nome = clean_val(row[11]) if len(row) >= 12 else None

                    esp_val = 'REVOLVER' if 'REVOLVER' in str(especie).upper() else 'PISTOLA'
                    vinc_val = 'PAISANO' if 'PAISANO' in str(vinculo).upper() else 'POLICIAL'
                    st_val = 'PARADO' if 'PARADO' in str(status).upper() else 'INICIADO'

                    ArmaTransferenciaPendente.objects.update_or_create(
                        numero_serie=serie,
                        nome_policial=nome,
                        defaults={
                            'especie': esp_val,
                            'marca': marca,
                            'modelo': modelo,
                            'calibre': calibre,
                            'tipo_vinculo': vinc_val,
                            're_policial': re_pol,
                            'situacao': situacao if situacao in ['FALECIDO', 'EXPULSO', 'BAIXA', 'TRANSFERENCIA'] else 'BAIXA',
                            'status': st_val,
                            'intencao_venda_nome': venda_nome
                        }
                    )
                    tr_count += 1
            except Exception as e:
                print(f"Erro em PST TRANSF: {e}")

        resumo['Transferências Pendentes'] = tr_count

    # -------------------------------------------------------------------------
    # 6. RÁDIOS HT (HT 26)
    # -------------------------------------------------------------------------
    if 'HT 26' in wb.sheetnames:
        ws = wb['HT 26']
        ht_count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                # Col C (index 2) is Motorola Serie e.g. 426CXK3307
                # Col D (index 3) is PMESP Patrimonio e.g. 221030854-P
                if len(row) >= 4 and row[2]:
                    serie = clean_val(row[2])
                    patrimonio = clean_val(row[3]) or f"HT-{serie}"
                    kit = clean_val(row[4])
                    sit = clean_val(row[5]) or 'OP'
                    dtic = clean_val(row[6])
                    data_dtic = parse_date(row[7]) if len(row) >= 8 else None

                    sit_val = 'OP'
                    if sit and 'MANUT' in str(sit).upper():
                        sit_val = 'MANUTENCAO'
                    elif sit and 'USO' in str(sit).upper():
                        sit_val = 'EM_USO'

                    if serie:
                        RadioHT.objects.update_or_create(
                            serie=serie,
                            defaults={
                                'patrimonio': patrimonio,
                                'kit_vinculado': str(kit) if kit else None,
                                'situacao': sit_val,
                                'chamado_dtic': dtic,
                                'data_chamado_dtic': data_dtic
                            }
                        )
                        ht_count += 1
            except Exception as e:
                print(f"Erro em Rádio HT: {e}")

        resumo['Rádios HT'] = ht_count

    # -------------------------------------------------------------------------
    # 7. AM 600, AM 640 e MOSQUETÃO (AM 600_640)
    # -------------------------------------------------------------------------
    if 'AM 600_640' in wb.sheetnames:
        ws = wb['AM 600_640']
        am640_c = 0
        am600_c = 0
        mosq_c = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            # AM 640 (Col C=serie, D=situacao)
            try:
                if len(row) >= 4 and row[2]:
                    serie = clean_val(row[2])
                    sit = clean_val(row[3]) or 'RESERVA'
                    if serie:
                        AM640.objects.update_or_create(
                            serie=serie,
                            defaults={'situacao': sit if sit in [c[0] for c in AM640.SITUACAO_CHOICES] else 'RESERVA'}
                        )
                        am640_c += 1
            except Exception as e:
                print(f"Erro em AM-640: {e}")

            # AM 600 (Col G=serie)
            try:
                if len(row) >= 8 and row[6]:
                    serie = clean_val(row[6])
                    if serie:
                        AM600.objects.update_or_create(serie=serie)
                        am600_c += 1
            except Exception as e:
                print(f"Erro em AM-600: {e}")

            # Mosquetão Federal (Col K=serie)
            try:
                if len(row) >= 12 and row[10]:
                    serie = clean_val(row[10])
                    if serie:
                        MosquetaoFederal.objects.update_or_create(serie=serie)
                        mosq_c += 1
            except Exception as e:
                print(f"Erro em Mosquetão: {e}")

        resumo['AM-640'] = am640_c
        resumo['AM-600'] = am600_c
        resumo['Mosquetão Federal'] = mosq_c

    # -------------------------------------------------------------------------
    # 8. TASER
    # -------------------------------------------------------------------------
    if 'TASER' in wb.sheetnames:
        ws = wb['TASER']
        tas_c = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                if len(row) >= 6 and row[3]:
                    serie = clean_val(row[3])
                    sit = clean_val(row[4]) or 'RESERVA'
                    carga = row[5]
                    try:
                        carga_int = int(float(carga)) if carga is not None else 100
                    except (ValueError, TypeError):
                        carga_int = 100

                    if serie:
                        TASER.objects.update_or_create(
                            serie=serie,
                            defaults={
                                'situacao': sit,
                                'carga_bateria_percent': max(0, min(100, carga_int))
                            }
                        )
                        tas_c += 1
            except Exception as e:
                print(f"Erro em TASER: {e}")

        resumo['TASER'] = tas_c

    # -------------------------------------------------------------------------
    # 9. ALGEMAS
    # -------------------------------------------------------------------------
    if 'ALGEMAS' in wb.sheetnames:
        ws = wb['ALGEMAS']
        alg_c = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                if len(row) >= 6 and row[5]:
                    num = clean_val(row[5])
                    if num:
                        Algemas.objects.update_or_create(numero=num)
                        alg_c += 1
            except Exception as e:
                print(f"Erro em Algemas: {e}")

        resumo['Algemas'] = alg_c

    # -------------------------------------------------------------------------
    # 10. MUNIÇÕES QUÍMICAS
    # -------------------------------------------------------------------------
    if 'QUÍMICAS' in wb.sheetnames:
        ws = wb['QUÍMICAS']
        mq_c = 0
        for row in ws.iter_rows(min_row=7, values_only=True):
            try:
                if len(row) >= 9 and row[2]:
                    desc = clean_val(row[2])
                    armario = row[3] or 0
                    kto = row[4] or 0
                    bornal = row[5] or 0
                    prazo = row[6]
                    vencidas = row[7] or 0

                    tipo_enum = None
                    for t_val, t_lbl in MunicaoQuimica.TIPO_CHOICES:
                        if desc and t_val.lower() in desc.lower():
                            tipo_enum = t_val
                            break
                    if not tipo_enum and desc:
                        tipo_enum = MunicaoQuimica.TIPO_CHOICES[0][0]

                    if desc and tipo_enum:
                        MunicaoQuimica.objects.update_or_create(
                            tipo_municao=tipo_enum,
                            defaults={
                                'qtd_armario': int(armario) if isinstance(armario, (int, float)) else 0,
                                'qtd_kto': int(kto) if isinstance(kto, (int, float)) else 0,
                                'qtd_bornal': int(bornal) if isinstance(bornal, (int, float)) else 0,
                                'qtd_vencidas': int(vencidas) if isinstance(vencidas, (int, float)) else 0,
                                'validade_prazo': parse_date(prazo),
                                'observacoes': desc
                            }
                        )
                        mq_c += 1
            except Exception as e:
                print(f"Erro em Munição Química: {e}")

        resumo['Munições Químicas'] = mq_c

    # -------------------------------------------------------------------------
    # 11. MUNIÇÕES CONVENCIONAIS
    # -------------------------------------------------------------------------
    if 'MUNIÇÕES ATUALIZADAS' in wb.sheetnames:
        ws = wb['MUNIÇÕES ATUALIZADAS']
        mc_c = 0
        current_calibre = '.40'
        for row in ws.iter_rows(min_row=1, values_only=True):
            try:
                r0 = str(row[0]) if row[0] is not None else ''
                if 'CAL .40' in r0.upper() or 'CAL 40' in r0.upper():
                    current_calibre = '.40'
                elif 'CAL. 556' in r0.upper() or 'CAL 556' in r0.upper():
                    current_calibre = '.556'
                elif 'CAL. 762' in r0.upper() or 'CAL 762' in r0.upper():
                    current_calibre = '.762'
                elif 'CAL. 12' in r0.upper() or 'CAL 12' in r0.upper():
                    current_calibre = '.12'

                if len(row) >= 7 and 'Cal' in r0:
                    subtipo_str = r0.split('-')[-1].strip().upper() if '-' in r0 else r0.strip()
                    subtipo_val = 'EXPO'
                    for st_code, st_lbl in MunicaoConvencional.SUBTIPO_CHOICES:
                        if st_code.upper() in subtipo_str or st_lbl.upper() in subtipo_str:
                            subtipo_val = st_code
                            break

                    em_uso = row[1] or 0
                    estoque = row[2] or 0
                    manuseadas = row[3] or 0
                    capsulas = row[4] or 0
                    danificado = row[5] or 0

                    MunicaoConvencional.objects.update_or_create(
                        calibre=current_calibre,
                        subtipo=subtipo_val,
                        secao='RESERVA',
                        defaults={
                            'em_uso': int(em_uso) if isinstance(em_uso, (int, float)) else 0,
                            'estoque': int(estoque) if isinstance(estoque, (int, float)) else 0,
                            'manuseadas': int(manuseadas) if isinstance(manuseadas, (int, float)) else 0,
                            'capsulas': int(capsulas) if isinstance(capsulas, (int, float)) else 0,
                            'danificado': int(danificado) if isinstance(danificado, (int, float)) else 0,
                        }
                    )
                    mc_c += 1
            except Exception as e:
                print(f"Erro em Munição Convencional: {e}")

        resumo['Munições Convencionais'] = mc_c

    # -------------------------------------------------------------------------
    # 12. COLETES BALÍSTICOS
    # -------------------------------------------------------------------------
    if 'CONTROLE DE COLETES' in wb.sheetnames:
        ws = wb['CONTROLE DE COLETES']
        col_c = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            try:
                if len(row) >= 5 and row[4]:
                    marca = clean_val(row[1]) or 'PROTECOP'
                    tamanho = clean_val(row[2]) or 'PADRÃO'
                    patrimonio = clean_val(row[3])
                    serie = clean_val(row[4])
                    situacao = clean_val(row[5]) or 'DISPONÍVEL'
                    obs = clean_val(row[6])
                    validade = clean_val(row[7]) or 'N/I'
                    capa = clean_val(row[8])

                    m_code = 'PROTECOP'
                    if 'KAVRO' in marca.upper():
                        m_code = 'KAVRO'
                    elif 'INBRATERRESTRE' in marca.upper():
                        m_code = 'INBRATERRESTRE'

                    sit_code = 'DISPONIVEL'
                    if situacao and 'SINDIC' in str(situacao).upper():
                        sit_code = 'SINDICANCIA'
                    elif situacao and 'USO' in str(situacao).upper():
                        sit_code = 'EM_USO'

                    tem_capa_val = True if (capa and str(capa).upper() in ('SIM', 'S', '1', 'X')) else False

                    if serie:
                        ColeteBalistico.objects.update_or_create(
                            numero_serie=serie,
                            defaults={
                                'marca': m_code,
                                'tamanho': tamanho,
                                'patrimonio': patrimonio,
                                'situacao': sit_code,
                                'validade_descricao': validade,
                                'tem_capa': tem_capa_val,
                                'obs': obs
                            }
                        )
                        col_c += 1
            except Exception as e:
                print(f"Erro em Colete: {e}")

        resumo['Coletes Balísticos'] = col_c

    # -------------------------------------------------------------------------
    # 13. ESCUDOS BALÍSTICOS
    # -------------------------------------------------------------------------
    if 'ESCUDOS BALÍSTICOS' in wb.sheetnames:
        ws = wb['ESCUDOS BALÍSTICOS']
        esc_c = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                if len(row) >= 3 and row[0] and isinstance(row[0], (int, float)):
                    num = int(row[0])
                    mat = clean_val(row[1]) or 'ESCUDO BALISTICO EM ARAMIDA / NIVEL I'
                    serie = clean_val(row[2])
                    fab = parse_date(row[3])
                    val = parse_date(row[4])
                    pat = clean_val(row[5])
                    qth = clean_val(row[6]) or 'RESERVA DE ARMAS'
                    lcm = clean_val(row[7])
                    sit = clean_val(row[8]) or 'OP'

                    lote_val = '1cia'
                    if lcm and '2' in str(lcm):
                        lote_val = '2cia'

                    sit_val = 'OP'
                    if sit and 'BX' in str(sit).upper():
                        sit_val = 'BXA'
                    elif sit and 'USO' in str(sit).upper():
                        sit_val = 'EM_USO'

                    EscudoBalistico.objects.update_or_create(
                        numero=num,
                        defaults={
                            'material': mat,
                            'numero_serie': serie,
                            'fabricacao': fab,
                            'validade': val,
                            'patrimonio': pat,
                            'localizacao': qth,
                            'lote_companhia': lote_val,
                            'situacao': sit_val
                        }
                    )
                    esc_c += 1
            except Exception as e:
                print(f"Erro em Escudo: {e}")

        resumo['Escudos Balísticos'] = esc_c

    # -------------------------------------------------------------------------
    # 14. CAPACETES BALÍSTICOS
    # -------------------------------------------------------------------------
    if 'CAPACETES BALÍSTICOS' in wb.sheetnames:
        ws = wb['CAPACETES BALÍSTICOS']
        cap_c = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            try:
                if len(row) >= 3 and row[0] and isinstance(row[0], (int, float)):
                    num = int(row[0])
                    mat = clean_val(row[1]) or 'CAP BAL NIVEL II C VISOR'
                    serie = clean_val(row[2])
                    pat = clean_val(row[3])
                    fab = parse_date(row[4])
                    val = clean_val(row[5]) or 'VENCIDO'
                    qth = clean_val(row[6]) or 'RESERVA DE ARMAS'
                    cond = clean_val(row[7]) or 'OPERANDO'
                    lcm = clean_val(row[8])

                    mat_code = 'COM_VISOR' if 'C VISOR' in str(mat).upper() else 'SEM_VISOR'
                    cond_code = 'OPERANDO'
                    if cond and 'DANIF' in str(cond).upper():
                        cond_code = 'DANIFICADO'
                    elif cond and 'USO' in str(cond).upper():
                        cond_code = 'EM_USO'

                    CapaceteBalistico.objects.update_or_create(
                        numero=num,
                        defaults={
                            'material': mat_code,
                            'numero_serie': serie,
                            'patrimonio': pat,
                            'fabricacao': fab,
                            'validade': val,
                            'localizacao': qth,
                            'condicao': cond_code,
                            'lote_companhia': lcm
                        }
                    )
                    cap_c += 1
            except Exception as e:
                print(f"Erro em Capacete: {e}")

        resumo['Capacetes Balísticos'] = cap_c

    # -------------------------------------------------------------------------
    # RESUMO FINAL
    # -------------------------------------------------------------------------
    print("==================================================")
    print("RESUMO DA IMPORTAÇÃO COMPLETA DA PLANILHA OFICIAL:")
    for k, v in resumo.items():
        print(f"  - {k}: {v} registros sincronizados com sucesso!")
    print("==================================================")


if __name__ == '__main__':
    run_import()

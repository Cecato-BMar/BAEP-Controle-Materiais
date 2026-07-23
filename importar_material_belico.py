"""
Importação da Planilha de Materiais Bélicos — 2º BAEP
Arquivo: PLANILHA DE MATERIAS - Atualizada .xlsx
Lê todas as abas e popula os models do módulo material_belico.
"""
import os, sys, django
sys.path.insert(0, os.getcwd())
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'reserva_baep.settings')
django.setup()

from openpyxl import load_workbook
from django.db import transaction
from material_belico.models import (
    Fuzil, EspingardaCal12, PistolaGlock, PistolaTaurus, ArmaTransferenciaPendente,
    RedDot, Magnificador, Supressor, VinculacaoAcessorioFuzil,
    KitOperacional, RadioHT, AM640, AM600, MosquetaoFederal,
    TASER, Algemas, MunicaoQuimica,
    MunicaoConvencional, DistribuicaoMunicaoKit,
    ColeteBalistico, EscudoBalistico, CapaceteBalistico,
)

FILE = 'PLANILHA DE MATERIAS - Atualizada .xlsx'

# =============================================================================
# HELPERS
# =============================================================================
def c(v):
    """Cell → stripped string, None-safe."""
    if v is None:
        return ''
    return str(v).strip()

def is_skip(v, extra=()):
    s = c(v).upper()
    if not s or s in ('', 'NAN', 'NONE', '────', '────────', '---'):
        return True
    if s in ('TOTAL', 'TOTAL =', 'ITEM', 'TIPO DE MATERIAL'):
        return True
    if s.startswith('FUZIS') or s.startswith('TOTAL '):
        return True
    if extra and s in extra:
        return True
    return False

def safe_int(v):
    try:
        return int(float(c(v)))
    except (ValueError, TypeError):
        return None

def normalize_status(raw):
    s = c(raw).upper()
    if 'SINDIC' in s:
        return 'SINDICANCIA'
    if 'BAIX' in s:
        return 'BAIXADO'
    if 'APREEND' in s:
        return 'APREENDIDA'
    if 'NOVIDADE' in s or 'NOVIDAD' in s:
        return 'NOVIDADE'
    return 'OK'

def normalize_loc(raw):
    s = c(raw).upper()
    valid = {
        'KIT-01','KIT-02','KIT-03','KIT-04','KIT-05','KIT-06',
        'KIT-07','KIT-08','KIT-09','KIT-10','KIT-11','KIT-12',
        'RESERVA','CMT','SUBCMT','AT-01','AT-02','AT-03','AT-04',
        'S-01','S-02','S-03','CURSO','P2','GUARDA',
    }
    if s in valid:
        return s
    if '3' in s and 'CIA' in s:
        return '3ª CIA'
    return s  # retorna como está; será tratado via observações

# =============================================================================
# 1) FUZIS E ACESSÓRIOS
# =============================================================================
def importar_fuzis(ws):
    """Importa Fuzis SCAR 762 e 556, Red Dot, Magnificador, Supressor, Vinculação."""
    stats = {'f762': 0, 'f556': 0, 'rd': 0, 'mag': 0, 'sup': 0, 'vinc': 0, 'err': 0}

    # --- SCAR CAL. 762 (cols C=3, D=4, E=5, F=6) rows 3..19 ---
    for r in range(3, 20):
        serial = c(ws.cell(r, 3).value)
        if is_skip(serial):
            continue
        loc = normalize_loc(ws.cell(r, 4).value)
        status = normalize_status(ws.cell(r, 6).value)
        num_rec = c(ws.cell(r, 5).value)
        try:
            Fuzil.objects.update_or_create(
                patrimonio=serial,
                defaults={'tipo': 'SCAR_762', 'localizacao': loc, 'status': status,
                          'numero_recibo': num_rec or None}
            )
            stats['f762'] += 1
        except Exception as e:
            print(f"  [ERRO] Fuzil 762 {serial}: {e}")
            stats['err'] += 1

    # --- SCAR CAL. 556 + IMBEL IA2 (cols I=9, J=10, K=11, L=12) rows 3..36 ---
    for r in range(3, 37):
        serial = c(ws.cell(r, 9).value)
        if is_skip(serial):
            continue
        loc = normalize_loc(ws.cell(r, 10).value)
        status = normalize_status(ws.cell(r, 12).value)
        num_rec = c(ws.cell(r, 11).value)
        # Determina tipo: SCAR_556 se serial começa com L/K, IMBEL_IA2 se começa com BRA/W
        if serial.upper().startswith(('BRA', 'W46', 'ESA')):
            tipo = 'IMBEL_IA2'
        else:
            tipo = 'SCAR_556'
        try:
            Fuzil.objects.update_or_create(
                patrimonio=serial,
                defaults={'tipo': tipo, 'localizacao': loc, 'status': status,
                          'numero_recibo': num_rec or None}
            )
            stats['f556'] += 1
        except Exception as e:
            print(f"  [ERRO] Fuzil 556 {serial}: {e}")
            stats['err'] += 1

    # --- RED DOT (cols O=15, P=16, Q=17) rows 3..57 ---
    for r in range(3, 58):
        pat = c(ws.cell(r, 15).value)
        if is_skip(pat):
            continue
        loc = normalize_loc(ws.cell(r, 16).value)
        status_raw = c(ws.cell(r, 17).value).upper()
        status = 'SINDICANCIA' if 'SINDIC' in status_raw else 'OK'
        try:
            RedDot.objects.update_or_create(
                patrimonio=pat,
                defaults={'localizacao': loc, 'status': status}
            )
            stats['rd'] += 1
        except Exception as e:
            print(f"  [ERRO] RedDot {pat}: {e}")
            stats['err'] += 1

    # --- MAGNIFICADOR (cols T=20, U=21, V=22) rows 3..28 ---
    for r in range(3, 29):
        pat = c(ws.cell(r, 20).value)
        if is_skip(pat):
            continue
        loc = normalize_loc(ws.cell(r, 21).value)
        status_raw = c(ws.cell(r, 22).value).upper()
        status = 'SINDICANCIA' if 'SINDIC' in status_raw else 'OK'
        try:
            Magnificador.objects.update_or_create(
                patrimonio=pat,
                defaults={'localizacao': loc, 'status': status}
            )
            stats['mag'] += 1
        except Exception as e:
            print(f"  [ERRO] Magnificador {pat}: {e}")
            stats['err'] += 1

    # --- SUPRESSOR (cols Y=25, Z=26) rows 3..6 ---
    for r in range(3, 7):
        pat = c(ws.cell(r, 25).value)
        if is_skip(pat):
            continue
        loc_raw = c(ws.cell(r, 26).value)
        # Supressores só aceitam AT-01..AT-04
        loc = loc_raw if loc_raw in ('AT-01', 'AT-02', 'AT-03', 'AT-04') else 'AT-01'
        try:
            Supressor.objects.update_or_create(
                patrimonio=pat,
                defaults={'localizacao': loc, 'status': 'OK'}
            )
            stats['sup'] += 1
        except Exception as e:
            print(f"  [ERRO] Supressor {pat}: {e}")
            stats['err'] += 1

    # --- VINCULAÇÃO (cols AH=34..AK=37) rows 3..38 ---
    for r in range(3, 39):
        fuzil_pat = c(ws.cell(r, 34).value)
        if is_skip(fuzil_pat):
            continue
        rd_pat = c(ws.cell(r, 35).value)
        mag_pat = c(ws.cell(r, 36).value)
        sup_pat = c(ws.cell(r, 37).value)
        try:
            fuzil = Fuzil.objects.filter(patrimonio=fuzil_pat).first()
            if not fuzil:
                continue
            rd = RedDot.objects.filter(patrimonio=rd_pat).first() if rd_pat and 'ACESS' not in rd_pat.upper() else None
            mag = Magnificador.objects.filter(patrimonio=mag_pat).first() if mag_pat and 'ACESS' not in mag_pat.upper() else None
            sup = Supressor.objects.filter(patrimonio=sup_pat).first() if sup_pat and 'ACESS' not in sup_pat.upper() else None
            VinculacaoAcessorioFuzil.objects.update_or_create(
                fuzil=fuzil,
                defaults={'red_dot': rd, 'magnificador': mag, 'supressor': sup}
            )
            stats['vinc'] += 1
        except Exception as e:
            print(f"  [ERRO] Vinculação {fuzil_pat}: {e}")
            stats['err'] += 1

    print(f"  Fuzis 762: {stats['f762']}  |  Fuzis 556/IA2: {stats['f556']}")
    print(f"  Red Dots: {stats['rd']}  |  Magnificadores: {stats['mag']}  |  Supressores: {stats['sup']}")
    print(f"  Vinculações: {stats['vinc']}  |  Erros: {stats['err']}")
    return stats


# =============================================================================
# 2) CAL.12 — ESPINGARDAS
# =============================================================================
def importar_cal12(ws):
    criados, erros = 0, 0
    for r in range(4, 27):  # rows 4..26
        num = c(ws.cell(r, 2).value)  # B: número
        if is_skip(num):
            continue
        pat = c(ws.cell(r, 3).value)  # C: patrimônio
        kit = c(ws.cell(r, 4).value)  # D: kit
        try:
            EspingardaCal12.objects.update_or_create(
                numero_espingarda=num,
                defaults={'patrimonio': pat or None, 'kit_vinculado': kit or None, 'status': 'OK'}
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Espingarda {num}: {e}")
            erros += 1
    print(f"  Espingardas Cal.12: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 3) PISTOLAS GLOCK
# =============================================================================
def importar_glock(ws):
    criados, erros = 0, 0
    for r in range(3, 31):  # rows 3..30
        serie = c(ws.cell(r, 3).value)  # C: Número/Série
        if is_skip(serie):
            continue
        pat = c(ws.cell(r, 2).value)  # B: Patrimônio
        modelo = c(ws.cell(r, 4).value) or 'PISTOLA GLOCK G22 G5 .40'
        cod_opm = c(ws.cell(r, 5).value) or None
        unidade = c(ws.cell(r, 6).value) or '2.BAEP'
        sit_raw = c(ws.cell(r, 7).value).upper()
        obs = c(ws.cell(r, 8).value) or None
        situacao = 'ok'
        bopm, bopc = None, None
        if 'APREEND' in sit_raw:
            situacao = 'APREENDIDA'
            # Extrai BOPM/BOPC das observações
            if obs:
                import re
                m_bopm = re.search(r'(?:BOPM\s*[Nnº°]?\s*)([\w/\-]+)', obs)
                m_bopc = re.search(r'(?:BOPC\s*)([\w/\-]+)', obs)
                bopm = m_bopm.group(1) if m_bopm else None
                bopc = m_bopc.group(1) if m_bopc else None
        elif 'NOVIDADE' in sit_raw:
            situacao = 'NOVIDADE'
        try:
            PistolaGlock.objects.update_or_create(
                numero_serie=serie,
                defaults={
                    'patrimonio': pat or None, 'modelo': modelo[:60],
                    'cod_opm': cod_opm, 'unidade': unidade[:20],
                    'situacao_reserva': situacao,
                    'numero_bopm': bopm, 'numero_bopc': bopc,
                    'observacoes': obs,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Glock {serie}: {e}")
            erros += 1
    print(f"  Pistolas Glock: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 4) PISTOLAS TAURUS
# =============================================================================
MODELO_TAURUS_MAP = {
    'PISTOLA TAURUS 24/7': 'TAURUS_24_7',
    'PISTOLA TAURUS PT100': 'TAURUS_PT100',
    'PISTOLA TAURUS 640': 'TAURUS_640',
    'PISTOLA TAURUS 641': 'TAURUS_640',  # 641 ≈ 640
}

def importar_taurus(ws):
    criados, erros = 0, 0
    for r in range(2, 32):  # rows 2..31
        serie = c(ws.cell(r, 3).value)  # C: Numero Serie
        if is_skip(serie):
            continue
        pat = c(ws.cell(r, 2).value)  # B: Patrimonio
        modelo_raw = c(ws.cell(r, 4).value)  # D: Modelo
        unidade = c(ws.cell(r, 5).value) or '2.BAEP'
        obs = c(ws.cell(r, 6).value) or None
        modelo_key = MODELO_TAURUS_MAP.get(modelo_raw, 'TAURUS_24_7')
        try:
            PistolaTaurus.objects.update_or_create(
                numero_serie=serie,
                defaults={
                    'patrimonio': pat or None, 'modelo': modelo_key,
                    'unidade': unidade[:20], 'observacoes': obs,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Taurus {serie}: {e}")
            erros += 1
    print(f"  Pistolas Taurus: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 5) PST TRANSF — TRANSFERÊNCIAS PENDENTES
# =============================================================================
def importar_transferencias(ws):
    criados, erros = 0, 0
    for r in range(2, 12):  # rows 2..11 (dados reais vão até ~6)
        especie = c(ws.cell(r, 2).value)  # B
        if not especie or especie.upper() in ('', 'NAN'):
            continue
        marca = c(ws.cell(r, 3).value)
        modelo = c(ws.cell(r, 4).value)
        calibre = c(ws.cell(r, 5).value)
        nserie = c(ws.cell(r, 6).value)
        nome = c(ws.cell(r, 7).value)
        tipo_vinc = c(ws.cell(r, 9).value).upper()
        re_pol = c(ws.cell(r, 10).value)
        sit_raw = c(ws.cell(r, 11).value).upper()
        status_raw = c(ws.cell(r, 12).value).upper()
        venda_nome = c(ws.cell(r, 13).value) or None
        venda_re = c(ws.cell(r, 14).value) or None
        # Mapeia espécie
        esp = 'PISTOLA' if 'PISTOLA' in especie.upper() else 'REVOLVER'
        tv = 'PAISANO' if 'PAISAN' in tipo_vinc else 'POLICIAL'
        sit = 'TRANSFERENCIA' if 'TRANSF' in sit_raw else sit_raw[:20]
        st = 'INICIADO' if 'INIC' in status_raw else 'PARADO'
        try:
            ArmaTransferenciaPendente.objects.update_or_create(
                numero_serie=nserie,
                defaults={
                    'especie': esp, 'marca': marca[:50], 'modelo': modelo[:100],
                    'calibre': calibre[:20], 'nome_policial': nome[:150],
                    'tipo_vinculo': tv, 're_policial': re_pol or None,
                    'situacao': sit, 'status': st,
                    'intencao_venda_nome': venda_nome, 'intencao_venda_re': venda_re,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Transferência {nserie}: {e}")
            erros += 1
    print(f"  Transferências Pendentes: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 6) AM 600/640 — MAGAZINES + MOSQUETÃO
# =============================================================================
def importar_am_mosquetao(ws):
    am640_c, am600_c, mosq_c, erros = 0, 0, 0, 0

    # AM 640: cols B=2, C=3, D=4, rows 3..26
    for r in range(3, 27):
        serie = c(ws.cell(r, 3).value)  # C: SÉRIE
        if is_skip(serie):
            continue
        sit_raw = c(ws.cell(r, 4).value).upper()  # D: SITUAÇÃO
        sit = sit_raw.replace(' ', '')
        # Normaliza KIT-XX
        if sit.startswith('KIT'):
            sit = sit.replace('KIT', 'KIT-').replace('KIT--', 'KIT-')
            if len(sit) == 5:
                sit = sit[:4] + '0' + sit[4:]
        try:
            AM640.objects.update_or_create(serie=serie, defaults={'situacao': sit[:20]})
            am640_c += 1
        except Exception as e:
            print(f"  [ERRO] AM640 {serie}: {e}")
            erros += 1

    # AM 600: cols F=6, G=7, H=8, rows 3..11
    for r in range(3, 12):
        serie = c(ws.cell(r, 7).value)  # G: SÉRIE
        if is_skip(serie):
            continue
        try:
            AM600.objects.update_or_create(serie=serie, defaults={})
            am600_c += 1
        except Exception as e:
            print(f"  [ERRO] AM600 {serie}: {e}")
            erros += 1

    # Mosquetão Federal: cols J=10, K=11, L=12, rows 3..4
    for r in range(3, 5):
        serie = c(ws.cell(r, 11).value)  # K: SÉRIE
        if is_skip(serie):
            continue
        try:
            MosquetaoFederal.objects.update_or_create(serie=serie, defaults={})
            mosq_c += 1
        except Exception as e:
            print(f"  [ERRO] Mosquetão {serie}: {e}")
            erros += 1

    print(f"  AM-640: {am640_c}  |  AM-600: {am600_c}  |  Mosquetão: {mosq_c}  |  Erros: {erros}")
    return am640_c + am600_c + mosq_c


# =============================================================================
# 7) HT 26 — RÁDIOS
# =============================================================================
def importar_radios(ws):
    criados, erros = 0, 0
    for r in range(3, 33):  # rows 3..32 (30 rádios)
        serie = c(ws.cell(r, 3).value)  # C: SÉRIE
        if is_skip(serie):
            continue
        pat = c(ws.cell(r, 2).value)  # B: PATRIMÔNIO
        kit_raw = c(ws.cell(r, 4).value)  # D: KIT
        sit_raw = c(ws.cell(r, 5).value).upper()  # E: SITUAÇÃO
        # Normaliza kit
        kit = kit_raw
        try:
            kit_int = int(float(kit_raw))
            kit = f'KIT-{kit_int:02d}'
        except (ValueError, TypeError):
            pass
        situacao = 'OP' if 'OP' in sit_raw or not sit_raw else 'MANUTENCAO'
        obs_parts = []
        if 'EMPRESTADO' in c(ws.cell(r, 4).value).upper() if ws.cell(r, 4).value else False:
            obs_parts.append('Emprestado da 3ª CIA')
        if 'EMPRESTADO' in kit_raw.upper():
            obs_parts.append('Emprestado da 3ª CIA')
            kit = kit_raw[:50]
        try:
            RadioHT.objects.update_or_create(
                serie=serie,
                defaults={
                    'patrimonio': pat or 'SEM-PAT', 'kit_vinculado': kit[:50],
                    'situacao': situacao, 'observacoes': '; '.join(obs_parts) or None,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Rádio {serie}: {e}")
            erros += 1
    print(f"  Rádios HT: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 8) TASER
# =============================================================================
def importar_taser(ws):
    criados, erros = 0, 0
    for r in range(3, 26):  # rows 3..25
        serie = c(ws.cell(r, 3).value)  # C: SÉRIE
        if is_skip(serie):
            continue
        sit = c(ws.cell(r, 4).value) or 'RESERVA'
        carga = safe_int(ws.cell(r, 5).value)
        if carga is None:
            carga = 0
        carga = max(0, min(100, carga))
        try:
            TASER.objects.update_or_create(
                serie=serie,
                defaults={'situacao': sit[:30], 'carga_bateria_percent': carga}
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] TASER {serie}: {e}")
            erros += 1
    print(f"  TASER: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 9) ALGEMAS
# =============================================================================
def importar_algemas(ws):
    criados, erros = 0, 0
    for r in range(3, 16):  # rows 3..15 (12 algemas)
        num = c(ws.cell(r, 4).value)  # D: número (PMESP XXXXX-XX)
        if is_skip(num):
            continue
        try:
            Algemas.objects.update_or_create(
                numero=num, defaults={'embalagem': 'PMESP'}
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Algemas {num}: {e}")
            erros += 1
    print(f"  Algemas: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 10) QUÍMICAS — MUNIÇÕES QUÍMICAS (tabela ARMAZENADAS NO ARMÁRIO)
# =============================================================================
TIPO_QUIM_MAP = {
    'GL-201': 'GL-201', 'GL-203L': 'GL-203L', 'GL-303': 'GL-303',
    'GL-304': 'GL-304', 'GL-307': 'GL-307', 'GL-300T': 'GL-300T',
    'GL-300TH': 'GL-300TH', 'GL CS 40MM': 'GL_CS_40mm',
    'GR HG CCS 60C': 'GR_HG_CCS_60C', 'AM-403P': 'AM-403P',
    'AM-403/P': 'AM-403P',
    'OC. AEROSOL': 'OC_AEROSOL', 'OC AEROSOL': 'OC_AEROSOL',
    'GRANADA FLASH BANG': 'FLASH_BANG', 'FLASH BANG': 'FLASH_BANG',
    'ESPARGIDOR PIMENTA 80': 'ESPARG_80',
    'ESPARGIDOR PIMENTA 400': 'ESPARG_400',
    '8909 NRSC': 'NRSC_GRANADA',
}

def importar_quimicas(ws):
    """Lê a tabela 'ARMAZENADAS NO ARMÁRIO' (~rows 41-57)."""
    criados, erros = 0, 0
    # ARMAZENADAS NO ARMÁRIO está na coluna B (dados começam ~row 33)
    # col B=2: ITEM, col C=3: DESCRIÇÃO, col D=4: ARMÁRIO, col G=7: VENCIDAS
    start_row = None
    for r in range(1, ws.max_row + 1):
        val = c(ws.cell(r, 2).value)
        if 'ARMAZENADAS NO ARMÁRIO' in val.upper():
            start_row = r + 4  # título + blank + header + sub-header
            break
    if not start_row:
        print("  [WARN] Tabela ARMAZENADAS NO ARMÁRIO não encontrada")
        return 0

    for r in range(start_row, start_row + 20):
        desc = c(ws.cell(r, 3).value)  # C: DESCRIÇÃO
        if not desc or desc.upper() in ('', 'NAN', 'ITEM'):
            continue
        armario = safe_int(ws.cell(r, 4).value) or 0
        vencidas = safe_int(ws.cell(r, 7).value) or 0
        # Mapeia tipo
        tipo_key = None
        desc_up = desc.upper()
        for prefix, key in TIPO_QUIM_MAP.items():
            if desc_up.startswith(prefix):
                tipo_key = key
                break
        if not tipo_key:
            continue
        try:
            MunicaoQuimica.objects.update_or_create(
                tipo_municao=tipo_key,
                defaults={'qtd_armario': armario, 'qtd_vencidas': vencidas}
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Munição Química {desc}: {e}")
            erros += 1
    print(f"  Munições Químicas: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 11) MUNIÇÕES CONVENCIONAIS
# =============================================================================
CAL_MAP = {'.40': '.40', 'CAL .40': '.40', 'CAL. 40': '.40',
           '.556': '.556', 'CAL. 556': '.556', 'CAL .556': '.556',
           '.762': '.762', 'CAL. 762': '.762', 'CAL .762': '.762',
           '.12': '.12', 'CAL.12': '.12', 'CAL .12': '.12'}
SUB_MAP = {'EXPO': 'EXPO', 'ETPP': 'ETPP', 'SS109': 'SS109', 'SAT': 'SAT',
           'TREINA': 'TREINA', 'TRAÇANTE': 'TRACANTE', 'TRACANTE': 'TRACANTE',
           'FESTIM': 'FESTIM', 'OP': 'OP', 'AP': 'AP',
           'BALOTE': 'BALOTE', 'SG': 'SG', '3T': '3T'}

def importar_municoes(ws):
    """Importa munições convencionais das seções RESERVA e CURSOS."""
    criados, erros = 0, 0

    # Seção RESERVA: rows ~3..22 (colunas B..G)
    # Seção CURSOS: rows ~31..45
    secoes = [('RESERVA', 3, 25), ('CURSOS', 29, 52)]

    for secao, r_start, r_end in secoes:
        for r in range(r_start, r_end):
            desc = c(ws.cell(r, 1).value)  # A: descrição (ex: "Cal .40 - EXPO")
            if not desc or 'MUNIÇÕES' in desc.upper() or 'QUADRO' in desc.upper():
                continue
            # Parse calibre e subtipo
            calibre, subtipo = None, None
            desc_up = desc.upper().replace('CAL.', 'CAL .').replace('  ', ' ')
            for prefix, cal in CAL_MAP.items():
                if desc_up.startswith(prefix.upper()):
                    calibre = cal
                    resto = desc_up.replace(prefix.upper(), '').replace(' - ', '').replace('-', '').strip()
                    for sub_key, sub_val in SUB_MAP.items():
                        if sub_key in resto:
                            subtipo = sub_val
                            break
                    break
            if not calibre or not subtipo:
                continue
            em_uso = safe_int(ws.cell(r, 2).value) or 0
            estoque = safe_int(ws.cell(r, 3).value) or 0
            manuseadas = safe_int(ws.cell(r, 4).value) or 0
            capsulas = safe_int(ws.cell(r, 5).value) or 0
            danificado = safe_int(ws.cell(r, 6).value) or 0
            try:
                MunicaoConvencional.objects.update_or_create(
                    calibre=calibre, subtipo=subtipo, secao=secao,
                    defaults={
                        'em_uso': em_uso, 'estoque': estoque,
                        'manuseadas': manuseadas, 'capsulas': capsulas,
                        'danificado': danificado,
                    }
                )
                criados += 1
            except Exception as e:
                print(f"  [ERRO] Munição {desc} ({secao}): {e}")
                erros += 1

    # Distribuição por Kit (cols L..O, rows 4..18)
    dist_c = 0
    for r in range(4, 19):
        kit_label = c(ws.cell(r, 13).value)  # M: "KIT 01", "CMT", etc
        if not kit_label or kit_label.upper() == 'TOTAL':
            continue
        # Mapeia kit_label -> numero_kit
        kit_num = kit_label.upper().replace('KIT ', 'KIT-').replace('KIT-0', 'KIT-')
        if kit_label.upper().startswith('KIT'):
            try:
                n = int(kit_label.split()[-1])
                kit_num = f'KIT-{n:02d}'
            except ValueError:
                kit_num = kit_label.upper()
        kit = KitOperacional.objects.filter(numero_kit=kit_num).first()
        if not kit:
            continue
        cal_556_ss = safe_int(ws.cell(r, 14).value) or 0  # N: Cal.556 SS109
        cal_556_sat = safe_int(ws.cell(r, 15).value) or 0  # O: Cal.556 SAT
        cal_762_op = safe_int(ws.cell(r, 16).value) or 0  # P: Cal.762 OP
        cal_762_ap = safe_int(ws.cell(r, 17).value) or 0  # Q: Cal.762 AP
        dists = [('.556', 'SS109', cal_556_ss), ('.556', 'SAT', cal_556_sat),
                 ('.762', 'OP', cal_762_op), ('.762', 'AP', cal_762_ap)]
        for cal, sub, qtd in dists:
            if qtd > 0:
                try:
                    DistribuicaoMunicaoKit.objects.update_or_create(
                        kit=kit, calibre=cal, subtipo=sub,
                        defaults={'quantidade_cota': qtd}
                    )
                    dist_c += 1
                except Exception:
                    pass

    # Distribuição por KTO (cols V..X, rows 4..18)
    for r in range(4, 19):
        kto_label = c(ws.cell(r, 19).value)  # S: "KTO 01", "CMT", etc
        if not kto_label or kto_label.upper() == 'TOTAL':
            continue
        kit_num = None
        if kto_label.upper().startswith('KTO'):
            try:
                n = int(kto_label.split()[-1])
                kit_num = f'KIT-{n:02d}'
            except ValueError:
                pass
        elif kto_label.upper() in ('CMT', 'SUBCMT'):
            kit_num = kto_label.upper()
        if not kit_num:
            continue
        kit = KitOperacional.objects.filter(numero_kit=kit_num).first()
        if not kit:
            continue
        cal_12_bal = safe_int(ws.cell(r, 20).value) or 0  # T: Cal.12 BALOTE
        cal_12_sg = safe_int(ws.cell(r, 21).value) or 0  # U: Cal.12 SG
        cal_12_3t = safe_int(ws.cell(r, 22).value) or 0  # V: Cal.12 3T
        dists2 = [('.12', 'BALOTE', cal_12_bal), ('.12', 'SG', cal_12_sg), ('.12', '3T', cal_12_3t)]
        for cal, sub, qtd in dists2:
            if qtd > 0:
                try:
                    DistribuicaoMunicaoKit.objects.update_or_create(
                        kit=kit, calibre=cal, subtipo=sub,
                        defaults={'quantidade_cota': qtd}
                    )
                    dist_c += 1
                except Exception:
                    pass

    print(f"  Munições Convencionais: {criados}  |  Distribuições: {dist_c}  |  Erros: {erros}")
    return criados + dist_c


# =============================================================================
# 12) CONTROLE DE COLETES
# =============================================================================
def importar_coletes(ws):
    criados, erros = 0, 0
    for r in range(3, 11):  # rows 3..10 (8 coletes)
        marca = c(ws.cell(r, 2).value)  # B
        if is_skip(marca):
            continue
        tamanho = c(ws.cell(r, 3).value)
        pat = c(ws.cell(r, 4).value) or None
        nserie = c(ws.cell(r, 5).value)
        sit_raw = c(ws.cell(r, 6).value).upper()
        obs = c(ws.cell(r, 7).value) or None
        validade = c(ws.cell(r, 8).value) or ''
        capa_raw = c(ws.cell(r, 9).value).upper()
        obs_ad = c(ws.cell(r, 10).value) or None
        situacao = 'SINDICANCIA' if 'SINDIC' in sit_raw else 'DISPONIVEL'
        tem_capa = capa_raw in ('SIM', 'S')
        # Parse ano e validade
        ano_fab, anos_val = None, None
        import re
        m = re.search(r'(\d+)\s*ANOS?\s*FAB\s*(\d{4})', validade.upper()) if validade else None
        if m:
            anos_val = int(m.group(1))
            ano_fab = int(m.group(2))
        try:
            ColeteBalistico.objects.update_or_create(
                numero_serie=nserie,
                defaults={
                    'marca': marca[:30], 'tamanho': tamanho[:30],
                    'patrimonio': pat, 'situacao': situacao,
                    'obs': obs, 'validade_descricao': validade[:50],
                    'ano_fabricacao': ano_fab, 'anos_validade': anos_val,
                    'tem_capa': tem_capa, 'obs_adicional': obs_ad,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Colete {nserie}: {e}")
            erros += 1
    print(f"  Coletes Balísticos: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 13) ESCUDOS BALÍSTICOS
# =============================================================================
def importar_escudos(ws):
    criados, erros = 0, 0
    for r in range(3, 66):  # rows 3..65 (63 escudos)
        num = safe_int(ws.cell(r, 1).value)  # A: Nº
        if num is None:
            continue
        material = c(ws.cell(r, 2).value) or 'ESCUDO BALISTICO EM ARAMIDA / NIVEL I'
        nserie = c(ws.cell(r, 3).value) or None
        fab_str = c(ws.cell(r, 4).value)
        val_str = c(ws.cell(r, 5).value)
        pat = c(ws.cell(r, 6).value) or None
        local_raw = c(ws.cell(r, 7).value) or 'RESERVA DE ARMAS'
        lote = c(ws.cell(r, 8).value) or None
        sit_raw = c(ws.cell(r, 9).value).upper()
        situacao = 'BXA' if 'BXA' in sit_raw or 'BX' in sit_raw else 'OP'
        # Parse dates
        from datetime import date as dt_date
        fab, val = None, None
        if fab_str and 'Invalid' not in fab_str:
            try:
                parts = fab_str.split('-')
                fab = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                pass
        if val_str and 'Invalid' not in val_str:
            try:
                parts = val_str.split('-')
                val = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                pass
        # Normaliza lote
        lote_norm = None
        if lote:
            if '1' in lote and 'CIA' in lote.upper():
                lote_norm = '1cia'
            elif '2' in lote and 'CIA' in lote.upper():
                lote_norm = '2cia'
            elif lote.upper() == 'EM':
                lote_norm = 'EM'
        try:
            EscudoBalistico.objects.update_or_create(
                numero=num,
                defaults={
                    'material': material[:100], 'numero_serie': nserie,
                    'fabricacao': fab, 'validade': val, 'patrimonio': pat,
                    'localizacao': local_raw[:50], 'lote_companhia': lote_norm,
                    'situacao': situacao,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Escudo {num}: {e}")
            erros += 1
    print(f"  Escudos Balísticos: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 14) CAPACETES BALÍSTICOS
# =============================================================================
def importar_capacetes(ws):
    criados, erros = 0, 0
    for r in range(3, 73):  # rows 3..72 (70 capacetes)
        num = safe_int(ws.cell(r, 1).value)  # A: Nº
        if num is None:
            continue
        material_raw = c(ws.cell(r, 2).value).upper()  # B: MATERIAL
        nserie = c(ws.cell(r, 3).value) or None
        pat = c(ws.cell(r, 4).value) or None
        validade = c(ws.cell(r, 6).value) or None  # F: VALIDADE
        local_raw = c(ws.cell(r, 7).value) or 'RESERVA DE ARMAS'  # G: QTH
        cond_raw = c(ws.cell(r, 8).value).upper()  # H: CONDIÇÃO
        lote = c(ws.cell(r, 9).value) or None  # I: LCM
        material = 'COM_VISOR' if 'VISOR' in material_raw or 'C VISOR' in material_raw else 'SEM_VISOR'
        condicao = 'DANIFICADO' if 'DANIF' in cond_raw or 'BXA' in cond_raw else 'OPERANDO'
        # Se situação é BXA (col J ou I), marca como DANIFICADO
        if lote and lote.upper() == 'BXA':
            condicao = 'DANIFICADO'
        try:
            CapaceteBalistico.objects.update_or_create(
                numero=num,
                defaults={
                    'material': material, 'numero_serie': nserie,
                    'patrimonio': pat, 'validade': validade[:30] if validade else None,
                    'localizacao': local_raw[:50], 'condicao': condicao,
                    'lote_companhia': lote[:30] if lote else None,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Capacete {num}: {e}")
            erros += 1
    print(f"  Capacetes Balísticos: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# 15) KITS OPERACIONAIS (a partir da aba KIT OP)
# =============================================================================
KIT_MAP = {
    'KIT OPERACIONAL': lambda n: f'KIT-{int(n):02d}',
    'KIT COMANDANTE': lambda n: 'CMT',
    'KIT SUBCOMANDANTE': lambda n: 'SUBCMT',
}

def importar_kits(ws):
    """Parse da aba KIT OP — kits lado a lado, blocos de 10 linhas."""
    criados, erros = 0, 0

    # Scan por blocos de kit: procura por "KIT OPERACIONAL" ou "KIT COMANDANTE" na coluna B
    kit_blocks = []
    for r in range(1, ws.max_row + 1):
        val = c(ws.cell(r, 2).value).upper()
        if 'KIT OPERACIONAL' in val or 'KIT COMANDANTE' in val or 'KIT SUBCOMANDANTE' in val:
            # Extrai número do kit
            import re
            m = re.search(r'(\d+)', val)
            if 'SUBCOMANDANTE' in val:
                kit_num = 'SUBCMT'
            elif 'COMANDANTE' in val and 'SUB' not in val:
                kit_num = 'CMT'
            elif m:
                kit_num = f'KIT-{int(m.group(1)):02d}'
            else:
                continue
            kit_blocks.append((r, kit_num, 'L'))  # lado esquerdo
            # Verifica lado direito (coluna G)
            val2 = c(ws.cell(r, 7).value).upper()
            if 'KIT' in val2:
                m2 = re.search(r'(\d+)', val2)
                if 'SUBCOMANDANTE' in val2:
                    kit_num2 = 'SUBCMT'
                elif 'COMANDANTE' in val2 and 'SUB' not in val2:
                    kit_num2 = 'CMT'
                elif m2:
                    kit_num2 = f'KIT-{int(m2.group(1)):02d}'
                else:
                    kit_num2 = None
                if kit_num2:
                    kit_blocks.append((r, kit_num2, 'R'))

    for start_row, kit_num, side in kit_blocks:
        # Colunas: L=B(2)E(5)  R=G(7)J(10)
        base = 2 if side == 'L' else 7
        data_start = start_row + 2  # cabeçalho + sub-cabeçalho

        # Coleta materiais do kit
        scar_556_pats = []
        imbel_pats = []
        scar_762_pat = None
        espingarda_pat = None
        ht_serie = None
        am640_serie = None
        escudo_num = None

        for i in range(8):  # 8 linhas de dados
            r = data_start + i
            tipo = c(ws.cell(r, base).value).upper()
            valor = c(ws.cell(r, base + 1).value)

            if 'SCAR' in tipo and '556' in tipo:
                if valor and not is_skip(valor):
                    scar_556_pats.append(valor)
            elif 'SCAR' in tipo and '762' in tipo:
                if valor and not is_skip(valor):
                    scar_762_pat = valor
            elif 'IMBEL' in tipo or 'IA2' in tipo:
                if valor and not is_skip(valor):
                    imbel_pats.append(valor)
            elif 'BENELLI' in tipo or 'CAL. 12' in tipo or 'CAL.12' in tipo:
                # Nº está na coluna valor (base+1)
                if valor and not is_skip(valor):
                    espingarda_pat = valor
            elif 'HT' in tipo:
                if valor and not is_skip(valor):
                    ht_serie = valor
            elif 'ESCUDO' in tipo:
                if valor and not is_skip(valor):
                    try:
                        escudo_num = int(float(valor))
                    except (ValueError, TypeError):
                        pass

        # Busca objetos no banco
        f556_1 = Fuzil.objects.filter(patrimonio__in=scar_556_pats).first() if scar_556_pats else None
        f556_2 = Fuzil.objects.filter(patrimonio__in=imbel_pats).first() if imbel_pats else None
        f762 = Fuzil.objects.filter(patrimonio=scar_762_pat).first() if scar_762_pat else None
        esp = EspingardaCal12.objects.filter(numero_espingarda=espingarda_pat).first() if espingarda_pat else None
        radio = RadioHT.objects.filter(serie=ht_serie).first() if ht_serie else None
        am = AM640.objects.filter(serie=am640_serie).first() if am640_serie else None
        esc = EscudoBalistico.objects.filter(numero=escudo_num).first() if escudo_num else None

        # Cria o kit
        obs_parts = []
        if len(scar_556_pats) > 1:
            obs_parts.append(f"SCAR 556 extras: {', '.join(scar_556_pats[1:])}")
        if len(imbel_pats) > 1:
            obs_parts.append(f"IMBEL extras: {', '.join(imbel_pats[1:])}")

        try:
            KitOperacional.objects.update_or_create(
                numero_kit=kit_num,
                defaults={
                    'fuzil_556_1': f556_1,
                    'fuzil_556_2': f556_2,
                    'fuzil_762': f762,
                    'espingarda': esp,
                    'radio_ht': radio,
                    'escudo': esc,
                    'observacoes': '; '.join(obs_parts) or None,
                }
            )
            criados += 1
        except Exception as e:
            print(f"  [ERRO] Kit {kit_num}: {e}")
            erros += 1

    print(f"  Kits Operacionais: {criados}  |  Erros: {erros}")
    return criados


# =============================================================================
# MAIN
# =============================================================================
if __name__ == '__main__':
    print("=" * 60)
    print("IMPORTAÇÃO — PLANILHA DE MATERIAIS BÉLICOS 2º BAEP")
    print("=" * 60)

    wb = load_workbook(FILE, data_only=True)
    sheets = wb.sheetnames
    print(f"Abas encontradas: {sheets}")

    def get_sheet(name):
        for s in sheets:
            if name.upper() in s.upper():
                return wb[s]
        return None

    print("\n[1/15] Fuzis e Acessórios...")
    ws = get_sheet('FUZIS')
    if ws:
        importar_fuzis(ws)
    else:
        print("  [WARN] Aba FUZIS E ACESSÓRIOS não encontrada!")

    print("\n[2/15] Espingardas Cal.12...")
    ws = get_sheet('CAL.12')
    if ws:
        importar_cal12(ws)
    else:
        print("  [WARN] Aba CAL.12 não encontrada!")

    print("\n[3/15] Pistolas Glock...")
    ws = get_sheet('GLOCK')
    if ws:
        importar_glock(ws)
    else:
        print("  [WARN] Aba PISTOLAS GLOCK não encontrada!")

    print("\n[4/15] Pistolas Taurus...")
    ws = get_sheet('TAURUS')
    if ws:
        importar_taurus(ws)
    else:
        print("  [WARN] Aba PISTOLAS TAURUS não encontrada!")

    print("\n[5/15] Transferências Pendentes...")
    ws = get_sheet('PST TRANSF')
    if ws:
        importar_transferencias(ws)
    else:
        print("  [WARN] Aba PST TRANSF não encontrada!")

    print("\n[6/15] AM-640, AM-600, Mosquetão...")
    ws = get_sheet('AM 600')
    if ws:
        importar_am_mosquetao(ws)
    else:
        print("  [WARN] Aba AM 600/640 não encontrada!")

    print("\n[7/15] Rádios HT...")
    ws = get_sheet('HT')
    if ws:
        importar_radios(ws)
    else:
        print("  [WARN] Aba HT 26 não encontrada!")

    print("\n[8/15] TASER...")
    ws = get_sheet('TASER')
    if ws:
        importar_taser(ws)
    else:
        print("  [WARN] Aba TASER não encontrada!")

    print("\n[9/15] Algemas...")
    ws = get_sheet('ALGEMAS')
    if ws:
        importar_algemas(ws)
    else:
        print("  [WARN] Aba ALGEMAS não encontrada!")

    print("\n[10/15] Munições Químicas...")
    ws = get_sheet('QUÍMICAS') or get_sheet('QUIMICAS')
    if ws:
        importar_quimicas(ws)
    else:
        print("  [WARN] Aba QUÍMICAS não encontrada!")

    print("\n[11/15] Coletes Balísticos...")
    ws = get_sheet('COLETES')
    if ws:
        importar_coletes(ws)
    else:
        print("  [WARN] Aba CONTROLE DE COLETES não encontrada!")

    print("\n[12/15] Escudos Balísticos...")
    ws = get_sheet('ESCUDOS')
    if ws:
        importar_escudos(ws)
    else:
        print("  [WARN] Aba ESCUDOS BALÍSTICOS não encontrada!")

    print("\n[13/15] Capacetes Balísticos...")
    ws = get_sheet('CAPACETES')
    if ws:
        importar_capacetes(ws)
    else:
        print("  [WARN] Aba CAPACETES BALÍSTICOS não encontrada!")

    print("\n[14/15] Kits Operacionais...")
    ws = get_sheet('KIT OP')
    if ws:
        importar_kits(ws)
    else:
        print("  [WARN] Aba KIT OP não encontrada!")

    # Munições DEPOIS dos kits (distribuição referencia KitOperacional)
    print("\n[15/15] Munições Convencionais + Distribuição...")
    ws = get_sheet('MUNIÇÕES') or get_sheet('MUNICOES')
    if ws:
        importar_municoes(ws)
    else:
        print("  [WARN] Aba MUNIÇÕES ATUALIZADAS não encontrada!")

    # Resumo final
    print("\n" + "=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    print(f"  Fuzis:              {Fuzil.objects.count()}")
    print(f"  Espingardas Cal.12: {EspingardaCal12.objects.count()}")
    print(f"  Pistolas Glock:     {PistolaGlock.objects.count()}")
    print(f"  Pistolas Taurus:    {PistolaTaurus.objects.count()}")
    print(f"  Transferências:     {ArmaTransferenciaPendente.objects.count()}")
    print(f"  Red Dots:           {RedDot.objects.count()}")
    print(f"  Magnificadores:     {Magnificador.objects.count()}")
    print(f"  Supressores:        {Supressor.objects.count()}")
    print(f"  Vinculações:        {VinculacaoAcessorioFuzil.objects.count()}")
    print(f"  Rádios HT:          {RadioHT.objects.count()}")
    print(f"  AM-640:             {AM640.objects.count()}")
    print(f"  AM-600:             {AM600.objects.count()}")
    print(f"  Mosquetões:         {MosquetaoFederal.objects.count()}")
    print(f"  TASERs:             {TASER.objects.count()}")
    print(f"  Algemas:            {Algemas.objects.count()}")
    print(f"  Munições Químicas:  {MunicaoQuimica.objects.count()}")
    print(f"  Munições Conv.:     {MunicaoConvencional.objects.count()}")
    print(f"  Distrib. Munição:   {DistribuicaoMunicaoKit.objects.count()}")
    print(f"  Coletes:            {ColeteBalistico.objects.count()}")
    print(f"  Escudos:            {EscudoBalistico.objects.count()}")
    print(f"  Capacetes:          {CapaceteBalistico.objects.count()}")
    print(f"  Kits Operacionais:  {KitOperacional.objects.count()}")
    print("=" * 60)
    print("IMPORTAÇÃO CONCLUÍDA!")
    print("=" * 60)

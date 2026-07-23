import datetime
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.db.models import Q, Count
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from movimentacoes.models import Movimentacao, Retirada, Devolucao
from materiais.models import Material
from policiais.models import Policial
from viaturas.models import DespachoViatura, Viatura

def is_admin_or_staff(user):
    """Verifica se o usuário é administrador ou membro da gestão."""
    return user.is_authenticated and (
        user.is_superuser or user.is_staff or 
        user.groups.filter(name__in=['administracao', 'reserva_armas', 'gestao']).exists()
    )

def parse_date(date_str, default=None):
    """Converte string YYYY-MM-DD ou DD/MM/YYYY para date object."""
    if not date_str:
        return default
    try:
        if '/' in date_str:
            d, m, y = date_str.split('/')
            return datetime.date(int(y), int(m), int(d))
        elif '-' in date_str:
            parts = date_str.split('-')
            return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        pass
    return default

@login_required
@user_passes_test(is_admin_or_staff)
def painel_administracao(request):
    """Painel de visão geral da administração do sistema."""
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    
    total_materiais = Material.objects.count()
    total_armas = Material.objects.filter(tipo='ARMA').count()
    armas_em_uso = Material.objects.filter(tipo='ARMA', status='EM_USO').count()
    movimentacoes_mes = Movimentacao.objects.filter(data_hora__date__gte=inicio_mes).count()
    policiais_ativos = Policial.objects.filter(situacao='ATIVO').count()
    
    context = {
        'total_materiais': total_materiais,
        'total_armas': total_armas,
        'armas_em_uso': armas_em_uso,
        'movimentacoes_mes': movimentacoes_mes,
        'policiais_ativos': policiais_ativos,
        'hoje': hoje,
    }
    return render(request, 'administracao/painel.html', context)

def obter_dados_consulta(params):
    """Função utilitária para extrair e filtrar dados de movimentação/cautelas e viaturas."""
    preset = params.get('preset', '')
    data_inicio_raw = params.get('data_inicio', '')
    data_fim_raw = params.get('data_fim', '')
    tipo_recurso = params.get('tipo_recurso', 'ARMA') # Padrão: ARMA
    categoria = params.get('categoria', '')
    status_devolucao = params.get('status_devolucao', 'TODOS')
    policial_id = params.get('policial_id', '')
    busca = params.get('busca', '').strip()

    hoje = timezone.now().date()
    
    # Tratamento de Presets
    if preset == 'hoje':
        data_inicio = hoje
        data_fim = hoje
    elif preset == 'ultimos_7_dias':
        data_inicio = hoje - datetime.timedelta(days=7)
        data_fim = hoje
    elif preset == 'este_mes':
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    elif preset == 'armas_julho_exemplo': # Preset do exemplo do usuário (10 a 14 de julho)
        ano = hoje.year
        data_inicio = datetime.date(ano, 7, 10)
        data_fim = datetime.date(ano, 7, 14)
    else:
        # Padrão: últimos 30 dias se nada for informado
        default_inicio = hoje - datetime.timedelta(days=30)
        data_inicio = parse_date(data_inicio_raw, default_inicio)
        data_fim = parse_date(data_fim_raw, hoje)

    # Garante intervalo com data_hora completa (start 00:00:00 até end 23:59:59)
    tz = timezone.get_current_timezone()
    dt_inicio = timezone.make_aware(datetime.datetime.combine(data_inicio, datetime.time.min), tz)
    dt_fim = timezone.make_aware(datetime.datetime.combine(data_fim, datetime.time.max), tz)

    itens_resultado = []
    estatisticas_categorias = {}
    total_registros = 0
    total_em_uso = 0
    total_devolvidos = 0

    if tipo_recurso in ['ARMA', 'MUNICAO', 'COLETE', 'RADIO', 'OUTROS', 'TODOS']:
        # Query de Movimentações de Retirada
        queryset = Movimentacao.objects.filter(
            tipo='RETIRADA',
            data_hora__range=(dt_inicio, dt_fim)
        ).select_related('material', 'policial', 'retirada', 'registrado_por').order_by('-data_hora')

        if tipo_recurso != 'TODOS':
            queryset = queryset.filter(material__tipo=tipo_recurso)
        
        if categoria:
            queryset = queryset.filter(material__categoria=categoria)

        if policial_id:
            queryset = queryset.filter(policial_id=policial_id)

        if busca:
            queryset = queryset.filter(
                Q(material__nome__icontains=busca) |
                Q(material__numero__icontains=busca) |
                Q(policial__nome__icontains=busca) |
                Q(policial__re__icontains=busca)
            )

        for mov in queryset:
            retirada = getattr(mov, 'retirada', None)
            devolucao_obj = Devolucao.objects.filter(retirada_referencia=retirada).select_related('movimentacao').first() if retirada else None
            
            is_devolvido = devolucao_obj is not None
            data_devolucao = devolucao_obj.movimentacao.data_hora if devolucao_obj else None

            if status_devolucao == 'DEVOLVIDO' and not is_devolvido:
                continue
            if status_devolucao == 'EM_USO' and is_devolvido:
                continue

            if is_devolvido:
                total_devolvidos += 1
            else:
                total_em_uso += 1

            cat_nome = mov.material.get_categoria_display() or mov.material.get_tipo_display()
            estatisticas_categorias[cat_nome] = estatisticas_categorias.get(cat_nome, 0) + mov.quantidade

            itens_resultado.append({
                'id': mov.id,
                'tipo_item': mov.material.get_tipo_display(),
                'categoria': cat_nome,
                'nome_material': mov.material.nome,
                'numero_serie': mov.material.numero,
                'quantidade': mov.quantidade,
                'policial_nome': str(mov.policial),
                'policial_rg': mov.policial.re,
                'policial_graduacao': mov.policial.get_posto_display() if hasattr(mov.policial, 'get_posto_display') else '',
                'data_retirada': mov.data_hora,
                'data_devolucao': data_devolucao,
                'status_devolucao': 'Devolvido' if is_devolvido else 'Em Uso / Cautelado',
                'is_devolvido': is_devolvido,
                'finalidade': retirada.finalidade if retirada else '',
                'local_uso': retirada.local_uso if retirada else '',
                'registrado_por': mov.registrado_por.get_full_name() or mov.registrado_por.username,
            })

    elif tipo_recurso == 'VIATURA':
        despachos = DespachoViatura.objects.filter(
            data_saida__range=(dt_inicio, dt_fim)
        ).select_related('viatura', 'motorista', 'encarregado', 'registrado_por').order_by('-data_saida')

        if policial_id:
            despachos = despachos.filter(Q(motorista_id=policial_id) | Q(encarregado_id=policial_id))

        if busca:
            despachos = despachos.filter(
                Q(viatura__prefixo__icontains=busca) |
                Q(viatura__placa__icontains=busca) |
                Q(motorista__nome__icontains=busca) |
                Q(motorista__re__icontains=busca)
            )

        for d in despachos:
            is_devolvido = d.data_retorno is not None
            if status_devolucao == 'DEVOLVIDO' and not is_devolvido:
                continue
            if status_devolucao == 'EM_USO' and is_devolvido:
                continue

            if is_devolvido:
                total_devolvidos += 1
            else:
                total_em_uso += 1

            cat_nome = d.viatura.modelo.nome if d.viatura.modelo else 'Viatura'
            estatisticas_categorias[cat_nome] = estatisticas_categorias.get(cat_nome, 0) + 1

            itens_resultado.append({
                'id': d.id,
                'tipo_item': 'Viatura',
                'categoria': cat_nome,
                'nome_material': f"Viatura Prefixo {d.viatura.prefixo}",
                'numero_serie': d.viatura.placa or d.viatura.prefixo,
                'quantidade': 1,
                'policial_nome': str(d.motorista),
                'policial_rg': d.motorista.re,
                'policial_graduacao': d.motorista.get_posto_display() if hasattr(d.motorista, 'get_posto_display') else '',
                'data_retirada': d.data_saida,
                'data_devolucao': d.data_retorno,
                'status_devolucao': 'Retornou' if is_devolvido else 'Em Patrulhamento / Uso',
                'is_devolvido': is_devolvido,
                'finalidade': f"Km Saída: {d.km_saida}" + (f" | Km Retorno: {d.km_retorno}" if d.km_retorno else ""),
                'local_uso': f"Encarregado: {d.encarregado}" if d.encarregado else '',
                'registrado_por': d.registrado_por.get_full_name() or d.registrado_por.username,
            })

    total_registros = len(itens_resultado)

    return {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'tipo_recurso': tipo_recurso,
        'categoria': categoria,
        'status_devolucao': status_devolucao,
        'policial_id': policial_id,
        'busca': busca,
        'preset': preset,
        'itens': itens_resultado,
        'estatisticas_categorias': estatisticas_categorias,
        'total_registros': total_registros,
        'total_em_uso': total_em_uso,
        'total_devolvidos': total_devolvidos,
    }

@login_required
@user_passes_test(is_admin_or_staff)
def consulta_dinamica(request):
    """Tela de pesquisa avançada periodizada e geração de relatório dinâmico."""
    dados = obter_dados_consulta(request.GET)
    
    # Listas auxiliares para dropdowns
    policiais = Policial.objects.filter(situacao='ATIVO').order_by('nome')
    categorias_arma = Material.CATEGORIA_CHOICES

    context = {
        **dados,
        'policiais': policiais,
        'categorias_arma': categorias_arma,
    }
    return render(request, 'administracao/consulta.html', context)

@login_required
@user_passes_test(is_admin_or_staff)
def imprimir_relatorio(request):
    """Layout formatado para impressão / visualização PDF oficial."""
    dados = obter_dados_consulta(request.GET)
    dados['agora'] = timezone.now()
    dados['usuario_solicitante'] = request.user.get_full_name() or request.user.username
    return render(request, 'administracao/relatorio_impressao.html', dados)

@login_required
@user_passes_test(is_admin_or_staff)
def exportar_excel(request):
    """Gera um arquivo .xlsx com o resultado do relatório filtrado."""
    dados = obter_dados_consulta(request.GET)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Utilização"

    # Estilos
    header_fill = PatternFill(start_color="1E272E", end_color="1E272E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="F1C40F")
    title_font = Font(name="Calibri", size=16, bold=True, color="1E272E")
    bold_font = Font(name="Calibri", size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Título do Relatório
    ws.merge_cells('A1:H1')
    ws['A1'] = "POLÍCIA MILITAR DO ESTADO DE SÃO PAULO - 2º BAEP"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:H2')
    data_ini_fmt = dados['data_inicio'].strftime('%d/%m/%Y')
    data_fim_fmt = dados['data_fim'].strftime('%d/%m/%Y')
    ws['A2'] = f"RELATÓRIO ADMINISTRATIVO DE UTILIZAÇÃO ({data_ini_fmt} a {data_fim_fmt})"
    ws['A2'].font = Font(name="Calibri", size=13, bold=True, color="2C3E50")
    ws['A2'].alignment = center_align

    ws.append([]) # Linha em branco

    # Resumo
    ws.append(["Resumo da Pesquisa:"])
    ws.cell(row=4, column=1).font = bold_font
    ws.append(["Total de Itens Utilizados:", dados['total_registros'], "", "Itens Devolvidos:", dados['total_devolvidos'], "", "Em Cautela/Uso:", dados['total_em_uso']])
    ws.append([])

    # Cabeçalho da Tabela
    headers = [
        "Item / Material", "Número / Série", "Tipo / Categoria", "Qtd",
        "Policial Responsável", "Data/Hora Retirada", "Data/Hora Devolução", "Status Atual"
    ]
    ws.append(headers)
    header_row_idx = ws.max_row
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Preenchimento de Dados
    for item in dados['itens']:
        dt_retirada_str = item['data_retirada'].strftime('%d/%m/%Y %H:%M') if item['data_retirada'] else '-'
        dt_devolucao_str = item['data_devolucao'].strftime('%d/%m/%Y %H:%M') if item['data_devolucao'] else 'Em Uso'
        
        row_data = [
            item['nome_material'],
            item['numero_serie'],
            item['categoria'],
            item['quantidade'],
            item['policial_nome'],
            dt_retirada_str,
            dt_devolucao_str,
            item['status_devolucao']
        ]
        ws.append(row_data)
        
        current_row = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.border = thin_border
            if col_idx in [2, 4, 6, 7, 8]:
                c.alignment = center_align

    # Ajuste automático de largura de colunas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Relatorio_Administrativo_{data_ini_fmt.replace('/','-')}_a_{data_fim_fmt.replace('/','-')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

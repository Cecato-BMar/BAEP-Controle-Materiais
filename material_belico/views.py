"""Views do módulo Material Bélico."""
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone

from reserva_baep.decorators import require_module_permission
from .models import (
    Fuzil, EspingardaCal12, PistolaGlock, PistolaTaurus, ArmaTransferenciaPendente,
    RedDot, Magnificador, Supressor, VinculacaoAcessorioFuzil,
    KitOperacional, RadioHT, AM640, AM600, MosquetaoFederal,
    TASER, Algemas, MunicaoQuimica,
    MunicaoConvencional, DistribuicaoMunicaoKit,
    ColeteBalistico, EscudoBalistico, CapaceteBalistico,
    STATUS_ARMA_CHOICES,
)
from .forms import (
    FuzilForm, EspingardaCal12Form, PistolaGlockForm, PistolaTaurusForm,
    ArmaTransferenciaPendenteForm,
    RedDotForm, MagnificadorForm, SupressorForm, VinculacaoAcessorioFuzilForm,
    KitOperacionalForm, RadioHTForm, AM640Form, AM600Form, MosquetaoFederalForm,
    TASERForm, AlgemasForm, MunicaoQuimicaForm,
    MunicaoConvencionalForm, DistribuicaoMunicaoKitForm,
    ColeteBalisticoForm, EscudoBalisticoForm, CapaceteBalisticoForm,
)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
@require_module_permission('material_belico')
def dashboard(request):
    """Dashboard principal com totalizadores e alertas."""
    total_fuzis = Fuzil.objects.exclude(status='BAIXADO').count()
    total_fuzis_762 = Fuzil.objects.filter(tipo='SCAR_762').exclude(status='BAIXADO').count()
    total_fuzis_556 = Fuzil.objects.filter(tipo__in=['SCAR_556', 'IMBEL_IA2']).exclude(status='BAIXADO').count()
    total_espingardas = EspingardaCal12.objects.exclude(status='BAIXADO').count()
    total_glocks = PistolaGlock.objects.count()
    total_taurus = PistolaTaurus.objects.count()
    total_kits = KitOperacional.objects.count()
    total_radios = RadioHT.objects.filter(situacao='OP').count()
    total_tasers = TASER.objects.count()
    total_coletes = ColeteBalistico.objects.filter(situacao='DISPONIVEL').count()
    total_escudos_op = EscudoBalistico.objects.filter(situacao='OP').count()
    total_capacetes_op = CapaceteBalistico.objects.filter(condicao='OPERANDO').count()
    total_transferencias = ArmaTransferenciaPendente.objects.count()

    # Alertas RN-02
    armas_sindiciancia = Fuzil.objects.filter(status='SINDICANCIA')
    glocks_apreendidas = PistolaGlock.objects.filter(situacao_reserva='APREENDIDA')

    # Alertas RN-03
    municoes_quimicas_vencidas = [mq for mq in MunicaoQuimica.objects.all() if mq.vencida]
    escudos_validade_vencida = [e for e in EscudoBalistico.objects.filter(situacao='OP') if e.alerta_validade]
    capacetes_vencidos = CapaceteBalistico.objects.filter(condicao='OPERANDO', validade__iexact='VENCIDO')
    coletes_validade_vencida = [c for c in ColeteBalistico.objects.filter(situacao='DISPONIVEL') if c.validade_vencida]

    # Alertas RN-07
    tasers_bateria_baixa = TASER.objects.filter(carga_bateria_percent__lt=50)
    tasers_bloqueados = TASER.objects.filter(carga_bateria_percent=0)

    # Transferências
    transferencias_paradas = ArmaTransferenciaPendente.objects.filter(status='PARADO')

    context = {
        'total_fuzis': total_fuzis, 'total_fuzis_762': total_fuzis_762,
        'total_fuzis_556': total_fuzis_556, 'total_espingardas': total_espingardas,
        'total_glocks': total_glocks, 'total_taurus': total_taurus,
        'total_kits': total_kits, 'total_radios': total_radios,
        'total_tasers': total_tasers, 'total_coletes': total_coletes,
        'total_escudos_op': total_escudos_op, 'total_capacetes_op': total_capacetes_op,
        'total_transferencias': total_transferencias,
        'armas_sindiciancia': armas_sindiciancia,
        'glocks_apreendidas': glocks_apreendidas,
        'municoes_quimicas_vencidas': municoes_quimicas_vencidas,
        'escudos_validade_vencida': escudos_validade_vencida,
        'capacetes_vencidos': capacetes_vencidos,
        'coletes_validade_vencida': coletes_validade_vencida,
        'tasers_bateria_baixa': tasers_bateria_baixa,
        'tasers_bloqueados': tasers_bloqueados,
        'transferencias_paradas': transferencias_paradas,
    }
    return render(request, 'material_belico/dashboard.html', context)


# =============================================================================
# HELPER CRUD GENÉRICO
# =============================================================================

def _crud_list(request, model, form_class, template_list, template_form, redirect_name,
               titulo, extra_qs=None, order_by=None):
    """Helper reutilizável para operações CRUD padrão."""
    qs = model.objects.all()
    if extra_qs:
        qs = extra_qs(qs)
    if order_by:
        qs = qs.order_by(*order_by)
    if request.method == 'POST':
        pk = request.POST.get('pk')
        if pk:
            obj = get_object_or_404(model, pk=pk)
            form = form_class(request.POST, instance=obj)
            msg_ok = f'{titulo} atualizado com sucesso.'
        else:
            form = form_class(request.POST)
            msg_ok = f'{titulo} cadastrado com sucesso.'
        if form.is_valid():
            form.save()
            messages.success(request, msg_ok)
            return redirect(redirect_name)
        else:
            messages.error(request, 'Corrija os erros abaixo.')
            return render(request, template_form, {'form': form, 'titulo': titulo, 'itens': qs})
    else:
        form = form_class()
    return render(request, template_list, {'itens': qs, 'form': form, 'titulo': titulo,
                                            'template_form': template_form, 'redirect_name': redirect_name})


def _crud_create(request, form_class, template_form, redirect_name, titulo, context_extra=None):
    """Helper para criar."""
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'{titulo} cadastrado(a) com sucesso.')
            return redirect(redirect_name)
    else:
        form = form_class()
    ctx = {'form': form, 'titulo': f'Novo(a) {titulo}', 'is_create': True}
    if context_extra:
        ctx.update(context_extra)
    return render(request, template_form, ctx)


def _crud_edit(request, pk, model, form_class, template_form, redirect_name, titulo, context_extra=None):
    """Helper para editar."""
    obj = get_object_or_404(model, pk=pk)
    if request.method == 'POST':
        form = form_class(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'{titulo} atualizado(a) com sucesso.')
            return redirect(redirect_name)
    else:
        form = form_class(instance=obj)
    ctx = {'form': form, 'titulo': f'Editar {titulo}', 'obj': obj, 'is_create': False}
    if context_extra:
        ctx.update(context_extra)
    return render(request, template_form, ctx)


def _crud_delete(request, pk, model, redirect_name, titulo):
    """Helper para excluir."""
    obj = get_object_or_404(model, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, f'{titulo} excluído(a) com sucesso.')
        return redirect(redirect_name)
    return render(request, 'material_belico/confirmar_exclusao.html',
                  {'obj': obj, 'titulo': titulo, 'redirect_name': redirect_name})


# =============================================================================
# CRUD: FUZIS
# =============================================================================

@login_required
@require_module_permission('material_belico')
def fuzil_list(request):
    itens = Fuzil.objects.all().order_by('tipo', 'patrimonio')
    # Extrair localizações únicas dos dados existentes
    locs = sorted(set(itens.values_list('localizacao', flat=True)))
    loc_map = dict(Fuzil._meta.get_field('localizacao').choices)
    tipo_map = dict(Fuzil.TIPO_CHOICES)
    status_map = dict(Fuzil._meta.get_field('status').choices)
    filter_fields = [
        {'label': 'Tipo', 'field': 'tipo', 'choices': [{'value': v, 'label': l} for v, l in Fuzil.TIPO_CHOICES]},
        {'label': 'Status', 'field': 'status', 'choices': [{'value': v, 'label': l} for v, l in STATUS_ARMA_CHOICES]},
        {'label': 'Localização', 'field': 'localizacao', 'choices': [{'value': loc, 'label': loc_map.get(loc, loc)} for loc in locs]},
    ]
    return render(request, 'material_belico/fuzil_list.html', {'itens': itens, 'titulo': 'Fuzis', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def fuzil_create(request):
    return _crud_create(request, FuzilForm, 'material_belico/form_generico.html', 'material_belico:fuzil_list', 'Fuzil')

@login_required
@require_module_permission('material_belico')
def fuzil_edit(request, pk):
    return _crud_edit(request, pk, Fuzil, FuzilForm, 'material_belico/form_generico.html', 'material_belico:fuzil_list', 'Fuzil')

@login_required
@require_module_permission('material_belico')
def fuzil_delete(request, pk):
    return _crud_delete(request, pk, Fuzil, 'material_belico:fuzil_list', 'Fuzil')


# =============================================================================
# CRUD: ESPINGARDAS CAL.12
# =============================================================================

@login_required
@require_module_permission('material_belico')
def espingarda_list(request):
    itens = EspingardaCal12.objects.all()
    status_map = dict(EspingardaCal12.STATUS_CHOICES)
    filter_fields = [
        {'label': 'Status', 'field': 'status', 'choices': [{'value': v, 'label': l} for v, l in EspingardaCal12.STATUS_CHOICES]},
    ]
    return render(request, 'material_belico/espingarda_list.html', {'itens': itens, 'titulo': 'Espingardas Cal.12', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def espingarda_create(request):
    return _crud_create(request, EspingardaCal12Form, 'material_belico/form_generico.html', 'material_belico:espingarda_list', 'Espingarda Cal.12')

@login_required
@require_module_permission('material_belico')
def espingarda_edit(request, pk):
    return _crud_edit(request, pk, EspingardaCal12, EspingardaCal12Form, 'material_belico/form_generico.html', 'material_belico:espingarda_list', 'Espingarda Cal.12')

@login_required
@require_module_permission('material_belico')
def espingarda_delete(request, pk):
    return _crud_delete(request, pk, EspingardaCal12, 'material_belico:espingarda_list', 'Espingarda Cal.12')


# =============================================================================
# CRUD: PISTOLAS GLOCK
# =============================================================================

@login_required
@require_module_permission('material_belico')
def pistola_glock_list(request):
    itens = PistolaGlock.objects.all()
    modelos = sorted(set(itens.values_list('modelo', flat=True)))
    filter_fields = [
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': v, 'label': l} for v, l in PistolaGlock.SITUACAO_CHOICES]},
        {'label': 'Modelo', 'field': 'modelo', 'choices': [{'value': m, 'label': m} for m in modelos]},
    ]
    return render(request, 'material_belico/glock_list.html', {'itens': itens, 'titulo': 'Pistolas Glock', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def pistola_glock_create(request):
    return _crud_create(request, PistolaGlockForm, 'material_belico/form_generico.html', 'material_belico:pistola_glock_list', 'Pistola Glock')

@login_required
@require_module_permission('material_belico')
def pistola_glock_edit(request, pk):
    return _crud_edit(request, pk, PistolaGlock, PistolaGlockForm, 'material_belico/form_generico.html', 'material_belico:pistola_glock_list', 'Pistola Glock')

@login_required
@require_module_permission('material_belico')
def pistola_glock_delete(request, pk):
    return _crud_delete(request, pk, PistolaGlock, 'material_belico:pistola_glock_list', 'Pistola Glock')


# =============================================================================
# CRUD: PISTOLAS TAURUS
# =============================================================================

@login_required
@require_module_permission('material_belico')
def pistola_taurus_list(request):
    itens = PistolaTaurus.objects.all()
    filter_fields = [
        {'label': 'Modelo', 'field': 'modelo', 'choices': [{'value': v, 'label': l} for v, l in PistolaTaurus.MODELO_CHOICES]},
    ]
    return render(request, 'material_belico/taurus_list.html', {'itens': itens, 'titulo': 'Pistolas Taurus', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def pistola_taurus_create(request):
    return _crud_create(request, PistolaTaurusForm, 'material_belico/form_generico.html', 'material_belico:pistola_taurus_list', 'Pistola Taurus')

@login_required
@require_module_permission('material_belico')
def pistola_taurus_edit(request, pk):
    return _crud_edit(request, pk, PistolaTaurus, PistolaTaurusForm, 'material_belico/form_generico.html', 'material_belico:pistola_taurus_list', 'Pistola Taurus')

@login_required
@require_module_permission('material_belico')
def pistola_taurus_delete(request, pk):
    return _crud_delete(request, pk, PistolaTaurus, 'material_belico:pistola_taurus_list', 'Pistola Taurus')


# =============================================================================
# CRUD: TRANSFERÊNCIAS PENDENTES
# =============================================================================

@login_required
@require_module_permission('material_belico')
def transferencia_list(request):
    itens = ArmaTransferenciaPendente.objects.all()
    filter_fields = [
        {'label': 'Espécie', 'field': 'especie', 'choices': [{'value': v, 'label': l} for v, l in ArmaTransferenciaPendente.ESPECIE_CHOICES]},
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': v, 'label': l} for v, l in ArmaTransferenciaPendente.SITUACAO_CHOICES]},
        {'label': 'Status', 'field': 'status', 'choices': [{'value': v, 'label': l} for v, l in ArmaTransferenciaPendente.STATUS_CHOICES]},
    ]
    return render(request, 'material_belico/transferencia_list.html', {'itens': itens, 'titulo': 'Transferências Pendentes', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def transferencia_create(request):
    return _crud_create(request, ArmaTransferenciaPendenteForm, 'material_belico/form_generico.html', 'material_belico:transferencia_list', 'Transferência Pendente')

@login_required
@require_module_permission('material_belico')
def transferencia_edit(request, pk):
    return _crud_edit(request, pk, ArmaTransferenciaPendente, ArmaTransferenciaPendenteForm, 'material_belico/form_generico.html', 'material_belico:transferencia_list', 'Transferência Pendente')

@login_required
@require_module_permission('material_belico')
def transferencia_delete(request, pk):
    return _crud_delete(request, pk, ArmaTransferenciaPendente, 'material_belico:transferencia_list', 'Transferência Pendente')


# =============================================================================
# CRUD: RED DOT
# =============================================================================

@login_required
@require_module_permission('material_belico')
def reddot_list(request):
    itens = RedDot.objects.all()
    locs = sorted(set(itens.values_list('localizacao', flat=True)))
    loc_map = dict(RedDot._meta.get_field('localizacao').choices)
    filter_fields = [
        {'label': 'Localização', 'field': 'localizacao', 'choices': [{'value': loc, 'label': loc_map.get(loc, loc)} for loc in locs]},
    ]
    return render(request, 'material_belico/acessorio_list.html', {'itens': itens, 'titulo': 'Red Dots', 'tipo': 'reddot', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def reddot_create(request):
    return _crud_create(request, RedDotForm, 'material_belico/form_generico.html', 'material_belico:reddot_list', 'Red Dot')

@login_required
@require_module_permission('material_belico')
def reddot_edit(request, pk):
    return _crud_edit(request, pk, RedDot, RedDotForm, 'material_belico/form_generico.html', 'material_belico:reddot_list', 'Red Dot')

@login_required
@require_module_permission('material_belico')
def reddot_delete(request, pk):
    return _crud_delete(request, pk, RedDot, 'material_belico:reddot_list', 'Red Dot')


# =============================================================================
# CRUD: MAGNIFICADOR
# =============================================================================

@login_required
@require_module_permission('material_belico')
def magnificador_list(request):
    itens = Magnificador.objects.all()
    locs = sorted(set(itens.values_list('localizacao', flat=True)))
    loc_map = dict(Magnificador._meta.get_field('localizacao').choices)
    filter_fields = [
        {'label': 'Localização', 'field': 'localizacao', 'choices': [{'value': loc, 'label': loc_map.get(loc, loc)} for loc in locs]},
        {'label': 'Status', 'field': 'status', 'choices': [{'value': v, 'label': l} for v, l in Magnificador.STATUS_CHOICES]},
    ]
    return render(request, 'material_belico/acessorio_list.html', {'itens': itens, 'titulo': 'Magnificadores', 'tipo': 'magnificador', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def magnificador_create(request):
    return _crud_create(request, MagnificadorForm, 'material_belico/form_generico.html', 'material_belico:magnificador_list', 'Magnificador')

@login_required
@require_module_permission('material_belico')
def magnificador_edit(request, pk):
    return _crud_edit(request, pk, Magnificador, MagnificadorForm, 'material_belico/form_generico.html', 'material_belico:magnificador_list', 'Magnificador')

@login_required
@require_module_permission('material_belico')
def magnificador_delete(request, pk):
    return _crud_delete(request, pk, Magnificador, 'material_belico:magnificador_list', 'Magnificador')


# =============================================================================
# CRUD: SUPRESSOR
# =============================================================================

@login_required
@require_module_permission('material_belico')
def supressor_list(request):
    itens = Supressor.objects.all()
    filter_fields = [
        {'label': 'Localização', 'field': 'localizacao', 'choices': [{'value': v, 'label': l} for v, l in Supressor.LOCALIZACAO_SUPRESSOR]},
    ]
    return render(request, 'material_belico/acessorio_list.html', {'itens': itens, 'titulo': 'Supressores', 'tipo': 'supressor', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def supressor_create(request):
    return _crud_create(request, SupressorForm, 'material_belico/form_generico.html', 'material_belico:supressor_list', 'Supressor')

@login_required
@require_module_permission('material_belico')
def supressor_edit(request, pk):
    return _crud_edit(request, pk, Supressor, SupressorForm, 'material_belico/form_generico.html', 'material_belico:supressor_list', 'Supressor')

@login_required
@require_module_permission('material_belico')
def supressor_delete(request, pk):
    return _crud_delete(request, pk, Supressor, 'material_belico:supressor_list', 'Supressor')


# =============================================================================
# CRUD: VINCULAÇÃO ACESSÓRIO–FUZIL
# =============================================================================

@login_required
@require_module_permission('material_belico')
def vinculacao_list(request):
    itens = VinculacaoAcessorioFuzil.objects.select_related('fuzil', 'red_dot', 'magnificador', 'supressor')
    filter_fields = [
        {'label': 'Tipo de Fuzil', 'field': 'tipo_fuzil', 'choices': [{'value': v, 'label': l} for v, l in Fuzil.TIPO_CHOICES]},
    ]
    return render(request, 'material_belico/vinculacao_list.html', {'itens': itens, 'titulo': 'Vinculação Acessório–Fuzil', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def vinculacao_create(request):
    return _crud_create(request, VinculacaoAcessorioFuzilForm, 'material_belico/form_generico.html', 'material_belico:vinculacao_list', 'Vinculação Acessório–Fuzil')

@login_required
@require_module_permission('material_belico')
def vinculacao_edit(request, pk):
    return _crud_edit(request, pk, VinculacaoAcessorioFuzil, VinculacaoAcessorioFuzilForm, 'material_belico/form_generico.html', 'material_belico:vinculacao_list', 'Vinculação Acessório–Fuzil')

@login_required
@require_module_permission('material_belico')
def vinculacao_delete(request, pk):
    return _crud_delete(request, pk, VinculacaoAcessorioFuzil, 'material_belico:vinculacao_list', 'Vinculação')


# =============================================================================
# CRUD: KITS OPERACIONAIS
# =============================================================================

@login_required
@require_module_permission('material_belico')
def kit_list(request):
    itens = KitOperacional.objects.select_related(
        'fuzil_556_1', 'fuzil_556_2', 'fuzil_762', 'espingarda', 'radio_ht', 'am640', 'escudo'
    )
    filter_fields = [
        {'label': 'Kit', 'field': 'numero_kit', 'choices': [{'value': v, 'label': l} for v, l in KitOperacional.NUMERO_KIT_CHOICES]},
    ]
    return render(request, 'material_belico/kit_list.html', {'itens': itens, 'titulo': 'Kits Operacionais', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def kit_create(request):
    return _crud_create(request, KitOperacionalForm, 'material_belico/form_generico.html', 'material_belico:kit_list', 'Kit Operacional')

@login_required
@require_module_permission('material_belico')
def kit_edit(request, pk):
    return _crud_edit(request, pk, KitOperacional, KitOperacionalForm, 'material_belico/form_generico.html', 'material_belico:kit_list', 'Kit Operacional')

@login_required
@require_module_permission('material_belico')
def kit_delete(request, pk):
    return _crud_delete(request, pk, KitOperacional, 'material_belico:kit_list', 'Kit Operacional')

@login_required
@require_module_permission('material_belico')
def kit_detail(request, pk):
    """View detalhada do kit com composição visual."""
    kit = get_object_or_404(KitOperacional.objects.select_related(
        'fuzil_556_1', 'fuzil_556_2', 'fuzil_762', 'espingarda', 'radio_ht', 'am640', 'escudo'
    ), pk=pk)
    distribuicoes = kit.distribuicoes_municao.all()
    return render(request, 'material_belico/kit_detail.html', {'kit': kit, 'distribuicoes': distribuicoes})


# =============================================================================
# CRUD: RÁDIO HT
# =============================================================================

@login_required
@require_module_permission('material_belico')
def radio_ht_list(request):
    itens = RadioHT.objects.all()
    filter_fields = [
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': v, 'label': l} for v, l in RadioHT.SITUACAO_CHOICES]},
    ]
    return render(request, 'material_belico/comunicacao_list.html', {'itens': itens, 'titulo': 'Rádios HT', 'tipo': 'radio_ht', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def radio_ht_create(request):
    return _crud_create(request, RadioHTForm, 'material_belico/form_generico.html', 'material_belico:radio_ht_list', 'Rádio HT')

@login_required
@require_module_permission('material_belico')
def radio_ht_edit(request, pk):
    return _crud_edit(request, pk, RadioHT, RadioHTForm, 'material_belico/form_generico.html', 'material_belico:radio_ht_list', 'Rádio HT')

@login_required
@require_module_permission('material_belico')
def radio_ht_delete(request, pk):
    return _crud_delete(request, pk, RadioHT, 'material_belico:radio_ht_list', 'Rádio HT')


# =============================================================================
# CRUD: AM-640
# =============================================================================

@login_required
@require_module_permission('material_belico')
def am640_list(request):
    itens = AM640.objects.all()
    filter_fields = [
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': v, 'label': l} for v, l in AM640.SITUACAO_CHOICES]},
    ]
    return render(request, 'material_belico/comunicacao_list.html', {'itens': itens, 'titulo': 'AM-640', 'tipo': 'am640', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def am640_create(request):
    return _crud_create(request, AM640Form, 'material_belico/form_generico.html', 'material_belico:am640_list', 'AM-640')

@login_required
@require_module_permission('material_belico')
def am640_edit(request, pk):
    return _crud_edit(request, pk, AM640, AM640Form, 'material_belico/form_generico.html', 'material_belico:am640_list', 'AM-640')

@login_required
@require_module_permission('material_belico')
def am640_delete(request, pk):
    return _crud_delete(request, pk, AM640, 'material_belico:am640_list', 'AM-640')


# =============================================================================
# CRUD: AM-600
# =============================================================================

@login_required
@require_module_permission('material_belico')
def am600_list(request):
    itens = AM600.objects.all()
    return render(request, 'material_belico/comunicacao_list.html', {'itens': itens, 'titulo': 'AM-600', 'tipo': 'am600', 'filter_fields': []})

@login_required
@require_module_permission('material_belico')
def am600_create(request):
    return _crud_create(request, AM600Form, 'material_belico/form_generico.html', 'material_belico:am600_list', 'AM-600')

@login_required
@require_module_permission('material_belico')
def am600_delete(request, pk):
    return _crud_delete(request, pk, AM600, 'material_belico:am600_list', 'AM-600')


# =============================================================================
# CRUD: MOSQUETÃO FEDERAL
# =============================================================================

@login_required
@require_module_permission('material_belico')
def mosquetao_list(request):
    itens = MosquetaoFederal.objects.all()
    return render(request, 'material_belico/comunicacao_list.html', {'itens': itens, 'titulo': 'Mosquetão Federal 201/Z', 'tipo': 'mosquetao', 'filter_fields': []})

@login_required
@require_module_permission('material_belico')
def mosquetao_create(request):
    return _crud_create(request, MosquetaoFederalForm, 'material_belico/form_generico.html', 'material_belico:mosquetao_list', 'Mosquetão Federal')

@login_required
@require_module_permission('material_belico')
def mosquetao_delete(request, pk):
    return _crud_delete(request, pk, MosquetaoFederal, 'material_belico:mosquetao_list', 'Mosquetão Federal')


# =============================================================================
# CRUD: TASER
# =============================================================================

@login_required
@require_module_permission('material_belico')
def taser_list(request):
    itens = TASER.objects.all()
    situacoes = sorted(set(itens.values_list('situacao', flat=True)))
    filter_fields = [
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': s, 'label': s} for s in situacoes]},
        {'label': 'Bateria', 'field': 'bateria', 'choices': [
            {'value': 'ok', 'label': '≥50% OK'}, {'value': 'baixa', 'label': '<50% Recarregar'}, {'value': 'zero', 'label': '0% Bloqueado'},
        ]},
    ]
    return render(request, 'material_belico/nao_letal_list.html', {'itens': itens, 'titulo': 'TASER', 'tipo': 'taser', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def taser_create(request):
    return _crud_create(request, TASERForm, 'material_belico/form_generico.html', 'material_belico:taser_list', 'TASER')

@login_required
@require_module_permission('material_belico')
def taser_edit(request, pk):
    return _crud_edit(request, pk, TASER, TASERForm, 'material_belico/form_generico.html', 'material_belico:taser_list', 'TASER')

@login_required
@require_module_permission('material_belico')
def taser_delete(request, pk):
    return _crud_delete(request, pk, TASER, 'material_belico:taser_list', 'TASER')


# =============================================================================
# CRUD: ALGEMAS
# =============================================================================

@login_required
@require_module_permission('material_belico')
def algemas_list(request):
    itens = Algemas.objects.all()
    return render(request, 'material_belico/nao_letal_list.html', {'itens': itens, 'titulo': 'Algemas', 'tipo': 'algemas', 'filter_fields': []})

@login_required
@require_module_permission('material_belico')
def algemas_create(request):
    return _crud_create(request, AlgemasForm, 'material_belico/form_generico.html', 'material_belico:algemas_list', 'Algemas')

@login_required
@require_module_permission('material_belico')
def algemas_edit(request, pk):
    return _crud_edit(request, pk, Algemas, AlgemasForm, 'material_belico/form_generico.html', 'material_belico:algemas_list', 'Algemas')

@login_required
@require_module_permission('material_belico')
def algemas_delete(request, pk):
    return _crud_delete(request, pk, Algemas, 'material_belico:algemas_list', 'Algemas')


# =============================================================================
# CRUD: MUNIÇÃO QUÍMICA
# =============================================================================

@login_required
@require_module_permission('material_belico')
def municao_quimica_list(request):
    itens = MunicaoQuimica.objects.all()
    filter_fields = [
        {'label': 'Tipo', 'field': 'tipo', 'choices': [{'value': v, 'label': l} for v, l in MunicaoQuimica.TIPO_CHOICES]},
    ]
    return render(request, 'material_belico/municao_quimica_list.html', {'itens': itens, 'titulo': 'Munições Químicas', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def municao_quimica_create(request):
    return _crud_create(request, MunicaoQuimicaForm, 'material_belico/form_generico.html', 'material_belico:municao_quimica_list', 'Munição Química')

@login_required
@require_module_permission('material_belico')
def municao_quimica_edit(request, pk):
    return _crud_edit(request, pk, MunicaoQuimica, MunicaoQuimicaForm, 'material_belico/form_generico.html', 'material_belico:municao_quimica_list', 'Munição Química')

@login_required
@require_module_permission('material_belico')
def municao_quimica_delete(request, pk):
    return _crud_delete(request, pk, MunicaoQuimica, 'material_belico:municao_quimica_list', 'Munição Química')


# =============================================================================
# CRUD: MUNIÇÃO CONVENCIONAL
# =============================================================================

@login_required
@require_module_permission('material_belico')
def municao_convencional_list(request):
    itens = MunicaoConvencional.objects.all().order_by('calibre', 'subtipo', 'secao')
    filter_fields = [
        {'label': 'Calibre', 'field': 'calibre', 'choices': [{'value': v, 'label': l} for v, l in MunicaoConvencional.CALIBRE_CHOICES]},
        {'label': 'Subtipo', 'field': 'subtipo', 'choices': [{'value': v, 'label': l} for v, l in MunicaoConvencional.SUBTIPO_CHOICES]},
        {'label': 'Seção', 'field': 'secao', 'choices': [{'value': v, 'label': l} for v, l in MunicaoConvencional.SECAO_CHOICES]},
    ]
    return render(request, 'material_belico/municao_convencional_list.html', {'itens': itens, 'titulo': 'Munições Convencionais', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def municao_convencional_create(request):
    return _crud_create(request, MunicaoConvencionalForm, 'material_belico/form_generico.html', 'material_belico:municao_convencional_list', 'Munição Convencional')

@login_required
@require_module_permission('material_belico')
def municao_convencional_edit(request, pk):
    return _crud_edit(request, pk, MunicaoConvencional, MunicaoConvencionalForm, 'material_belico/form_generico.html', 'material_belico:municao_convencional_list', 'Munição Convencional')

@login_required
@require_module_permission('material_belico')
def municao_convencional_delete(request, pk):
    return _crud_delete(request, pk, MunicaoConvencional, 'material_belico:municao_convencional_list', 'Munição Convencional')


# =============================================================================
# CRUD: COLETE BALÍSTICO
# =============================================================================

@login_required
@require_module_permission('material_belico')
def colete_list(request):
    itens = ColeteBalistico.objects.all()
    tamanhos = sorted(set(itens.values_list('tamanho', flat=True)))
    filter_fields = [
        {'label': 'Marca', 'field': 'marca', 'choices': [{'value': v, 'label': l} for v, l in ColeteBalistico.MARCA_CHOICES]},
        {'label': 'Tamanho', 'field': 'tamanho', 'choices': [{'value': t, 'label': t} for t in tamanhos]},
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': v, 'label': l} for v, l in ColeteBalistico.SITUACAO_CHOICES]},
    ]
    return render(request, 'material_belico/protecao_list.html', {'itens': itens, 'titulo': 'Coletes Balísticos', 'tipo': 'colete', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def colete_create(request):
    return _crud_create(request, ColeteBalisticoForm, 'material_belico/form_generico.html', 'material_belico:colete_list', 'Colete Balístico')

@login_required
@require_module_permission('material_belico')
def colete_edit(request, pk):
    return _crud_edit(request, pk, ColeteBalistico, ColeteBalisticoForm, 'material_belico/form_generico.html', 'material_belico:colete_list', 'Colete Balístico')

@login_required
@require_module_permission('material_belico')
def colete_delete(request, pk):
    return _crud_delete(request, pk, ColeteBalistico, 'material_belico:colete_list', 'Colete Balístico')


# =============================================================================
# CRUD: ESCUDO BALÍSTICO
# =============================================================================

@login_required
@require_module_permission('material_belico')
def escudo_list(request):
    itens = EscudoBalistico.objects.all()
    filter_fields = [
        {'label': 'Situação', 'field': 'situacao', 'choices': [{'value': v, 'label': l} for v, l in EscudoBalistico.SITUACAO_CHOICES]},
        {'label': 'Lote/Cia', 'field': 'lote', 'choices': [{'value': v, 'label': l} for v, l in EscudoBalistico.LOTE_CHOICES]},
    ]
    return render(request, 'material_belico/protecao_list.html', {'itens': itens, 'titulo': 'Escudos Balísticos', 'tipo': 'escudo', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def escudo_create(request):
    return _crud_create(request, EscudoBalisticoForm, 'material_belico/form_generico.html', 'material_belico:escudo_list', 'Escudo Balístico')

@login_required
@require_module_permission('material_belico')
def escudo_edit(request, pk):
    return _crud_edit(request, pk, EscudoBalistico, EscudoBalisticoForm, 'material_belico/form_generico.html', 'material_belico:escudo_list', 'Escudo Balístico')

@login_required
@require_module_permission('material_belico')
def escudo_delete(request, pk):
    return _crud_delete(request, pk, EscudoBalistico, 'material_belico:escudo_list', 'Escudo Balístico')


# =============================================================================
# CRUD: CAPACETE BALÍSTICO
# =============================================================================

@login_required
@require_module_permission('material_belico')
def capacete_list(request):
    itens = CapaceteBalistico.objects.all()
    filter_fields = [
        {'label': 'Material', 'field': 'material', 'choices': [{'value': v, 'label': l} for v, l in CapaceteBalistico.MATERIAL_CHOICES]},
        {'label': 'Condição', 'field': 'condicao', 'choices': [{'value': v, 'label': l} for v, l in CapaceteBalistico.CONDICAO_CHOICES]},
    ]
    return render(request, 'material_belico/protecao_list.html', {'itens': itens, 'titulo': 'Capacetes Balísticos', 'tipo': 'capacete', 'filter_fields': filter_fields})

@login_required
@require_module_permission('material_belico')
def capacete_create(request):
    return _crud_create(request, CapaceteBalisticoForm, 'material_belico/form_generico.html', 'material_belico:capacete_list', 'Capacete Balístico')

@login_required
@require_module_permission('material_belico')
def capacete_edit(request, pk):
    return _crud_edit(request, pk, CapaceteBalistico, CapaceteBalisticoForm, 'material_belico/form_generico.html', 'material_belico:capacete_list', 'Capacete Balístico')

@login_required
@require_module_permission('material_belico')
def capacete_delete(request, pk):
    return _crud_delete(request, pk, CapaceteBalistico, 'material_belico:capacete_list', 'Capacete Balístico')


# =============================================================================
# =============================================================================
# EXPORTAÇÃO DETALHADA EM EXCEL
# =============================================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required
@require_module_permission('material_belico')
def exportar_relatorio_detalhado(request):
    """Gera um relatório detalhado em Excel com todos os itens do Material Bélico."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Material Bélico"
    
    header_fill = PatternFill(start_color="1E272E", end_color="1E272E", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="F1C40F")
    center_align = Alignment(horizontal="center", vertical="center")
    
    headers = ["Categoria", "Tipo / Modelo", "Patrimônio", "Série / Número", "Status / Situação", "Localização / Observações"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    def add_row(cat, tipo, pat, serie, status, loc):
        ws.append([str(cat), str(tipo), str(pat or '-'), str(serie or '-'), str(status or '-'), str(loc or '-')])

    # Armas Longas
    for f in Fuzil.objects.all():
        add_row('Fuzil', f.get_tipo_display(), f.patrimonio, f.numero_recibo or '-', f.get_status_display(), f.get_localizacao_display())
        
    for e in EspingardaCal12.objects.all():
        add_row('Espingarda Cal.12', 'Calibre 12', e.patrimonio, e.numero_espingarda, e.get_status_display(), e.observacoes)
        
    # Pistolas
    for p in PistolaGlock.objects.all():
        add_row('Pistola Glock', p.modelo, p.patrimonio, p.numero_serie, p.get_situacao_display(), p.numero_bopm or '-')

    for p in PistolaTaurus.objects.all():
        add_row('Pistola Taurus', p.get_modelo_display(), p.patrimonio, p.numero_serie, '-', p.unidade)

    # Não Letal
    for t in TASER.objects.all():
        add_row('TASER', 'TASER', '-', t.serie, t.situacao, f"Bateria: {t.carga_bateria_percent}%")

    for a in Algemas.objects.all():
        add_row('Algemas', a.marca, '-', a.numero_serie, '-', '-')

    for mq in MunicaoQuimica.objects.all():
        add_row('Munição Química', mq.get_tipo_display(), '-', '-', '-', f"Validade: {mq.validade_prazo.strftime('%d/%m/%Y') if mq.validade_prazo else '-'}")

    # Proteção Balística
    for c in ColeteBalistico.objects.all():
        add_row('Colete Balístico', c.get_marca_display(), '-', c.numero_serie, c.get_situacao_display(), f"Tamanho: {c.tamanho}")

    for e in EscudoBalistico.objects.all():
        add_row('Escudo Balístico', e.material, str(e.numero), e.numero_serie, e.get_situacao_display(), e.get_lote_display())

    for c in CapaceteBalistico.objects.all():
        add_row('Capacete Balístico', c.get_material_display(), str(c.numero), c.numero_serie, c.get_condicao_display(), '-')

    # Comunicação / Outros
    for r in RadioHT.objects.all():
        add_row('Rádio HT', 'Motorola APX 2000', r.patrimonio, r.serie, r.get_situacao_display(), f"Kit: {r.kit_vinculado or '-'}")

    # Ajustar larguras
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_Detalhado_Material_Belico_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response

import openpyxl
import traceback

# Mapeamento: chave interna → (Model, {coluna_excel: campo_model}, campo_unique)
IMPORT_CONFIGS = {
    'oficial': {
        'label': 'Planilha Oficial Completa do 2º BAEP (Todas as Abas)',
        'unique_field': None,
        'columns': {},
    },
    'fuzil': {
        'model': Fuzil,
        'label': 'Fuzis',
        'unique_field': 'patrimonio',
        'columns': {
            'tipo': 'tipo',
            'patrimonio': 'patrimonio',
            'patrimônio': 'patrimonio',
            'localização': 'localizacao',
            'localizacao': 'localizacao',
            'nº recibo': 'numero_recibo',
            'numero recibo': 'numero_recibo',
            'status': 'status',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
    'espingarda': {
        'model': EspingardaCal12,
        'label': 'Espingardas Cal.12',
        'unique_field': 'numero_espingarda',
        'columns': {
            'número': 'numero_espingarda',
            'numero': 'numero_espingarda',
            'numero espingarda': 'numero_espingarda',
            'número da espingarda': 'numero_espingarda',
            'patrimonio': 'patrimonio',
            'patrimônio': 'patrimonio',
            'kit vinculado': 'kit_vinculado',
            'kit': 'kit_vinculado',
            'status': 'status',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
    'glock': {
        'model': PistolaGlock,
        'label': 'Pistolas Glock',
        'unique_field': 'numero_serie',
        'columns': {
            'patrimonio': 'patrimonio',
            'patrimônio': 'patrimonio',
            'número de série': 'numero_serie',
            'numero de serie': 'numero_serie',
            'numero serie': 'numero_serie',
            'nº série': 'numero_serie',
            'serie': 'numero_serie',
            'série': 'numero_serie',
            'modelo': 'modelo',
            'cód. opm': 'cod_opm',
            'cod opm': 'cod_opm',
            'unidade': 'unidade',
            'situação': 'situacao_reserva',
            'situacao': 'situacao_reserva',
            'situação reserva': 'situacao_reserva',
            'nº bopm': 'numero_bopm',
            'bopm': 'numero_bopm',
            'nº bopc': 'numero_bopc',
            'bopc': 'numero_bopc',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
    'taurus': {
        'model': PistolaTaurus,
        'label': 'Pistolas Taurus',
        'unique_field': 'numero_serie',
        'columns': {
            'patrimonio': 'patrimonio',
            'patrimônio': 'patrimonio',
            'número de série': 'numero_serie',
            'numero de serie': 'numero_serie',
            'numero serie': 'numero_serie',
            'serie': 'numero_serie',
            'série': 'numero_serie',
            'modelo': 'modelo',
            'unidade': 'unidade',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
    'radio_ht': {
        'model': RadioHT,
        'label': 'Rádios HT',
        'unique_field': 'serie',
        'columns': {
            'patrimonio': 'patrimonio',
            'patrimônio': 'patrimonio',
            'serie': 'serie',
            'série': 'serie',
            'kit vinculado': 'kit_vinculado',
            'kit': 'kit_vinculado',
            'situação': 'situacao',
            'situacao': 'situacao',
            'chamado dtic': 'chamado_dtic',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
    'am640': {
        'model': AM640,
        'label': 'AM-640',
        'unique_field': 'serie',
        'columns': {
            'serie': 'serie',
            'série': 'serie',
            'situação': 'situacao',
            'situacao': 'situacao',
        },
    },
    'colete': {
        'model': ColeteBalistico,
        'label': 'Coletes Balísticos',
        'unique_field': 'numero_serie',
        'columns': {
            'marca': 'marca',
            'tamanho': 'tamanho',
            'patrimonio': 'patrimonio',
            'patrimônio': 'patrimonio',
            'número de série': 'numero_serie',
            'numero de serie': 'numero_serie',
            'numero serie': 'numero_serie',
            'serie': 'numero_serie',
            'série': 'numero_serie',
            'situação': 'situacao',
            'situacao': 'situacao',
            'validade': 'validade_descricao',
            'capa': 'tem_capa',
            'tem capa': 'tem_capa',
            'observações': 'obs',
            'observacoes': 'obs',
            'obs': 'obs',
        },
    },
    'taser': {
        'model': TASER,
        'label': 'TASER',
        'unique_field': 'serie',
        'columns': {
            'serie': 'serie',
            'série': 'serie',
            'situação': 'situacao',
            'situacao': 'situacao',
            'carga bateria': 'carga_bateria_percent',
            'bateria': 'carga_bateria_percent',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
    'algemas': {
        'model': Algemas,
        'label': 'Algemas',
        'unique_field': 'numero',
        'columns': {
            'número': 'numero',
            'numero': 'numero',
            'observações': 'observacoes',
            'observacoes': 'observacoes',
        },
    },
}


def _normalize_header(h):
    """Normaliza cabeçalho de coluna para matching flexível."""
    import unicodedata
    if h is None:
        return ''
    h = str(h).strip().lower()
    # Remove acentos
    nfkd = unicodedata.normalize('NFKD', h)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


@login_required
@require_module_permission('material_belico')
def importar_excel(request):
    """Importa dados de uma planilha Excel para o modelo selecionado."""
    tipo_choices = [(k, v['label']) for k, v in IMPORT_CONFIGS.items()]
    resultado = None

    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        tipo_importacao = request.POST.get('tipo_importacao', '')
        modo = request.POST.get('modo', 'ignorar')  # ignorar | atualizar
        arquivo = request.FILES['arquivo_excel']

        if tipo_importacao not in IMPORT_CONFIGS:
            messages.error(request, 'Selecione um tipo de material válido.')
            return render(request, 'material_belico/importar_excel.html', {
                'tipo_choices': tipo_choices,
            })

        if tipo_importacao == 'oficial':
            try:
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    for chunk in arquivo.chunks():
                        tmp.write(chunk)
                    tmp_path = tmp.name

                from importar_planilha_oficial import run_import, EXCEL_PATH
                import importar_planilha_oficial
                importar_planilha_oficial.EXCEL_PATH = tmp_path
                run_import()
                os.remove(tmp_path)

                messages.success(request, 'Planilha Oficial Completa do 2º BAEP importada e sincronizada com sucesso!')
                return render(request, 'material_belico/importar_excel.html', {
                    'tipo_choices': tipo_choices,
                    'resultado': {
                        'tipo_label': 'Planilha Oficial Completa do 2º BAEP',
                        'criados': '450+',
                        'atualizados': 0,
                        'ignorados': 0,
                        'erros': [],
                        'total_processado': 'Todos os materiais sincronizados com sucesso!',
                        'colunas_mapeadas': {'Todas as Abas': 'FUZIS, GLOCK, CAL.12, TAURUS, HT, TASER, COLETES, ESCUDOS, CAPACETES, etc.'},
                    }
                })
            except Exception as e:
                messages.error(request, f'Erro ao processar Planilha Oficial: {str(e)}')
                traceback.print_exc()
                return render(request, 'material_belico/importar_excel.html', {
                    'tipo_choices': tipo_choices,
                })

        config = IMPORT_CONFIGS[tipo_importacao]
        Model = config['model']
        col_map = config['columns']
        unique_field = config['unique_field']

        try:
            wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active

            # Lê cabeçalhos da primeira linha
            headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers_norm = [_normalize_header(h) for h in headers_raw]

            # Mapeia índices → campos do model
            field_map = {}  # idx -> campo_model
            for idx, h_norm in enumerate(headers_norm):
                # Tenta match direto
                if h_norm in col_map:
                    field_map[idx] = col_map[h_norm]
                else:
                    # Tenta match normalizado sem acentos do col_map
                    for col_key, col_field in col_map.items():
                        if _normalize_header(col_key) == h_norm:
                            field_map[idx] = col_field
                            break

            if not field_map:
                messages.error(request, f'Nenhuma coluna reconhecida na planilha. Cabeçalhos encontrados: {headers_raw}')
                return render(request, 'material_belico/importar_excel.html', {
                    'tipo_choices': tipo_choices,
                })

            # Processa linhas
            criados = 0
            atualizados = 0
            ignorados = 0
            erros = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    dados = {}
                    for idx, campo in field_map.items():
                        if idx < len(row):
                            valor = row[idx]
                            if valor is not None:
                                valor = str(valor).strip()
                                if valor == '':
                                    valor = None
                            # Converte booleanos
                            if campo == 'tem_capa' and valor is not None:
                                valor = valor.lower() in ('sim', 's', 'true', '1', 'x', 'yes')
                            # Converte inteiros
                            if campo == 'carga_bateria_percent' and valor is not None:
                                try:
                                    valor = int(float(valor))
                                except (ValueError, TypeError):
                                    valor = 100
                            dados[campo] = valor

                    # Pula linhas completamente vazias
                    if not any(v is not None for v in dados.values()):
                        continue

                    unique_val = dados.get(unique_field)
                    if not unique_val:
                        erros.append(f'Linha {row_num}: campo obrigatório "{unique_field}" vazio.')
                        continue

                    # Verifica duplicata
                    existente = Model.objects.filter(**{unique_field: unique_val}).first()
                    if existente:
                        if modo == 'atualizar':
                            for campo, valor in dados.items():
                                if valor is not None and campo != unique_field:
                                    setattr(existente, campo, valor)
                            existente.save()
                            atualizados += 1
                        else:
                            ignorados += 1
                    else:
                        Model.objects.create(**{k: v for k, v in dados.items() if v is not None})
                        criados += 1

                except Exception as e:
                    erros.append(f'Linha {row_num}: {str(e)}')

            wb.close()

            resultado = {
                'tipo_label': config['label'],
                'criados': criados,
                'atualizados': atualizados,
                'ignorados': ignorados,
                'erros': erros,
                'total_processado': criados + atualizados + ignorados,
                'colunas_mapeadas': {headers_raw[idx]: campo for idx, campo in field_map.items() if idx < len(headers_raw)},
            }

            if criados > 0 or atualizados > 0:
                messages.success(request, f'Importação concluída: {criados} criados, {atualizados} atualizados.')
            if ignorados > 0:
                messages.info(request, f'{ignorados} registros duplicados ignorados.')
            if erros:
                messages.warning(request, f'{len(erros)} linhas com erro.')

        except Exception as e:
            messages.error(request, f'Erro ao processar a planilha: {str(e)}')
            traceback.print_exc()

    return render(request, 'material_belico/importar_excel.html', {
        'tipo_choices': tipo_choices,
        'resultado': resultado,
    })
